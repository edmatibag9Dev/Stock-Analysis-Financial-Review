"""
PLTR & NOW Investment Analysis Model Builder
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = Workbook()

# Colors
BLUE   = "FF0000FF"
BLACK  = "FF000000"
GREEN  = "FF008000"
WHITE  = "FFFFFFFF"
DARK_NAVY  = "FF1F3864"
MID_BLUE   = "FF2E75B6"
LIGHT_BLUE = "FFD6E4F0"
TEAL       = "FF1F6B75"
LIGHT_TEAL = "FFD9EDF0"
GOLD       = "FFFFC000"
LIGHT_GOLD = "FFFFF2CC"
GREY_HDR   = "FF404040"
LIGHT_GREY = "FFF2F2F2"
MED_GREY   = "FFD9D9D9"

ARIAL = "Arial"

def font(color=BLACK, bold=False, size=10, italic=False):
    return Font(name=ARIAL, color=color, bold=bold, size=size, italic=italic)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def center(): return Alignment(horizontal="center", vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

FMT_CURRENCY = '$#,##0;($#,##0);"-"'
FMT_CURRENCY_DEC = '$#,##0.0;($#,##0.0);"-"'
FMT_PCT     = '0.0%;(0.0%);"-"'
FMT_MULT    = '0.0x;(0.0x);"-"'
FMT_NUM     = '#,##0;(#,##0);"-"'
FMT_TEXT    = '@'
FMT_PRICE   = '$#,##0.00;($#,##0.00);"-"'
FMT_RO40    = '#,##0.0;(#,##0.0);"-"'

def add_comment(cell, text):
    c = Comment(text, "Model")
    c.width = 280; c.height = 80
    cell.comment = c

def style_section_header(ws, row, c1, c2, label, bg=MID_BLUE):
    ws.cell(row, c1).value = label
    for c in range(c1, c2+1):
        cell = ws.cell(row, c)
        cell.font = font(WHITE, bold=True)
        cell.fill = fill(bg)
        cell.alignment = center() if c > c1 else left()

def style_header_row(ws, row, c1, c2, bg=DARK_NAVY):
    for c in range(c1, c2+1):
        cell = ws.cell(row, c)
        cell.font = font(WHITE, bold=True)
        cell.fill = fill(bg)
        cell.alignment = center()

def set_col_widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

years = ["FY2026","FY2027","FY2028","FY2029","FY2030"]
scenarios = ["Bull","Base","Bear"]
scenario_colors = {"Bull": TEAL, "Base": MID_BLUE, "Bear": GREY_HDR}
scenario_light  = {"Bull": LIGHT_TEAL, "Base": LIGHT_BLUE, "Bear": LIGHT_GREY}

# ════════════════════════════════════════════════════════
# SHEET 2: PLTR_Model
# ════════════════════════════════════════════════════════
ws_pltr = wb.active
ws_pltr.title = "PLTR_Model"
set_col_widths(ws_pltr, {"A":36,"B":14,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14})
freeze(ws_pltr)

r = 1
ws_pltr.merge_cells(f"A{r}:J{r}")
c = ws_pltr.cell(r,1,"PALANTIR TECHNOLOGIES (PLTR) — Investment Analysis Model")
c.font = font(WHITE, bold=True, size=13); c.fill = fill(DARK_NAVY); c.alignment = center()
ws_pltr.row_dimensions[r].height = 28

# Section A
r += 1
style_section_header(ws_pltr, r, 1, 10, "SECTION A — HISTORICAL DATA (FY2024–Q1 2026)")
r += 1

hdrs = ["Metric","FY2024","FY2025","Q1 2025","Q2 2025","Q3 2025","Q4 2025","Q1 2026"]
for ci, h in enumerate(hdrs, 1):
    ws_pltr.cell(r,ci).value = h
style_header_row(ws_pltr, r, 1, 10)
r += 1

pltr_hist = [
    ("Revenue ($M)",         [2865,4477,883.9,1004,1181,1408,1632.6], FMT_CURRENCY,
     "Source: Palantir Q4 2025 & Q1 2026 Earnings Press Releases, May 2026"),
    ("YoY Revenue Growth",   [None,0.563,None,None,None,None,0.869], FMT_PCT,
     "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
    ("Gross Margin",         [None,None,None,None,None,None,0.87], FMT_PCT,
     "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
    ("Adj. Operating Margin",[0.36,0.50,None,None,None,None,0.60], FMT_PCT,
     "Source: Palantir FY2024-FY2025 Annual Reports; Q1 2026 Press Release"),
    ("GAAP Operating Margin",[None,None,None,None,None,None,0.46], FMT_PCT,
     "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
    ("Adj. FCF Margin",      [None,None,None,None,None,None,0.57], FMT_PCT,
     "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
    ("Adj. FCF ($M)",        [None,2239,None,None,None,None,None], FMT_CURRENCY,
     "Source: Palantir FY2025 Annual Report (estimate)"),
    ("Adj. EBITDA ($M)",     [None,None,None,None,None,None,990.3], FMT_CURRENCY,
     "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
    ("Adj. EBITDA Margin",   [None,None,None,None,None,None,0.61], FMT_PCT,
     "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
    ("SBC ($M)",             [None,None,None,None,None,None,201.6], FMT_CURRENCY,
     "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
    ("SBC % of Revenue",     [None,None,None,None,None,None,0.123], FMT_PCT,
     "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
]

for label, vals, fmt, src in pltr_hist:
    ws_pltr.cell(r,1).value = label
    ws_pltr.cell(r,1).font = font(BLACK)
    ws_pltr.cell(r,1).alignment = left()
    if r % 2 == 0: ws_pltr.cell(r,1).fill = fill(LIGHT_GREY)
    for ci, v in enumerate(vals, 2):
        cell = ws_pltr.cell(r,ci)
        if v is not None:
            cell.value = v; cell.font = font(BLUE)
            add_comment(cell, src)
        cell.number_format = fmt; cell.alignment = center()
        if r % 2 == 0: cell.fill = fill(LIGHT_GREY)
    r += 1

# Segment detail
r += 1
style_section_header(ws_pltr, r, 1, 10, "Revenue Segment Detail (Q1 2026)", MID_BLUE)
r += 1
for ci, h in enumerate(["Metric","Q1 2026 ($M)","YoY Growth"], 1):
    ws_pltr.cell(r,ci).value = h
style_header_row(ws_pltr, r, 1, 10)
r += 1

segs = [
    ("US Revenue ($M)",            1282, 1.04, "Source: Palantir Q1 2026 Press Release, May 4 2026"),
    ("US Commercial Revenue ($M)", 595,  1.33, "Source: Palantir Q1 2026 Press Release, May 4 2026"),
    ("US Government Revenue ($M)", 687,  0.84, "Source: Palantir Q1 2026 Press Release, May 4 2026"),
    ("International Revenue ($M)", 351,  None, "Source: Palantir Q1 2026 Press Release, May 4 2026"),
    ("Total Customers (TTM)",      1007, None, "Source: Palantir Q1 2026 Press Release, May 4 2026"),
]
for label, val, gr, src in segs:
    ws_pltr.cell(r,1).value = label; ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    cell_v = ws_pltr.cell(r,2)
    cell_v.value = val; cell_v.font = font(BLUE); cell_v.alignment = center()
    cell_v.number_format = FMT_NUM if "Customer" in label else FMT_CURRENCY
    add_comment(cell_v, src)
    if gr is not None:
        cell_g = ws_pltr.cell(r,3)
        cell_g.value = gr; cell_g.font = font(BLUE); cell_g.alignment = center()
        cell_g.number_format = FMT_PCT
        add_comment(cell_g, src)
    if r % 2 == 0:
        for c in range(1,11): ws_pltr.cell(r,c).fill = fill(LIGHT_GREY)
    r += 1

# Section B: Projections
r += 2
style_section_header(ws_pltr, r, 1, 10, "SECTION B — REVENUE & FCF PROJECTIONS (FY2026–FY2030)", DARK_NAVY)
r += 1

ws_pltr.cell(r,1).value = "FY2025 Revenue ($M) [Base year for projections]"
ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
ws_pltr.cell(r,2).value = 4477; ws_pltr.cell(r,2).font = font(BLUE)
ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
add_comment(ws_pltr.cell(r,2), "Source: Palantir FY2025 Earnings / Q4 2025 Press Release")
pltr_fy25_rev_row = r
r += 1

ws_pltr.cell(r,1).value = "Year"
for ci, yr in enumerate(years, 2): ws_pltr.cell(r,ci).value = yr
style_header_row(ws_pltr, r, 1, 7)
r += 1

pltr_growth_rates = {
    "Bull": [0.71, 0.60, 0.50, 0.40, 0.30],
    "Base": [0.71, 0.50, 0.35, 0.25, 0.18],
    "Bear": [0.71, 0.35, 0.20, 0.15, 0.10],
}
pltr_fcf_margins = {
    "Bull": [0.58, 0.60, 0.62, 0.62, 0.62],
    "Base": [0.56, 0.56, 0.55, 0.54, 0.52],
    "Bear": [0.52, 0.50, 0.47, 0.44, 0.40],
}

pltr_growth_rows = {}
pltr_fcf_margin_rows = {}
pltr_rev_rows = {}
pltr_fcf_rows = {}

for scen in scenarios:
    style_section_header(ws_pltr, r, 1, 10, f"{scen} Scenario", scenario_colors[scen])
    r += 1

    ws_pltr.cell(r,1).value = f"{scen} Rev Growth Assumption"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    pltr_growth_rows[scen] = r
    src_g = f"Source: Palantir FY2026 guidance 71% YoY confirmed; {scen} scenario analyst estimates"
    for ci, g in enumerate(pltr_growth_rates[scen], 2):
        cell = ws_pltr.cell(r,ci)
        cell.value = g; cell.font = font(BLUE)
        cell.number_format = FMT_PCT; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
        add_comment(cell, src_g)
    r += 1

    ws_pltr.cell(r,1).value = f"{scen} Revenue ($M)"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    pltr_rev_rows[scen] = r
    for ci in range(2, 7):
        col = get_column_letter(ci)
        cell = ws_pltr.cell(r,ci)
        if ci == 2:
            cell.value = f"=$B${pltr_fy25_rev_row}*(1+B{pltr_growth_rows[scen]})"
        else:
            prev = get_column_letter(ci-1)
            cell.value = f"={prev}{r}*(1+{col}{pltr_growth_rows[scen]})"
        cell.font = font(BLACK); cell.number_format = FMT_CURRENCY; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
    r += 1

    ws_pltr.cell(r,1).value = f"{scen} FCF Margin Assumption"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    pltr_fcf_margin_rows[scen] = r
    src_m = f"Source: PLTR Q1 2026 adj FCF margin 57%; {scen} scenario analyst estimates"
    for ci, m in enumerate(pltr_fcf_margins[scen], 2):
        cell = ws_pltr.cell(r,ci)
        cell.value = m; cell.font = font(BLUE)
        cell.number_format = FMT_PCT; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
        add_comment(cell, src_m)
    r += 1

    ws_pltr.cell(r,1).value = f"{scen} FCF ($M)"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    pltr_fcf_rows[scen] = r
    for ci in range(2, 7):
        col = get_column_letter(ci)
        cell = ws_pltr.cell(r,ci)
        cell.value = f"={col}{pltr_rev_rows[scen]}*{col}{pltr_fcf_margin_rows[scen]}"
        cell.font = font(BLACK); cell.number_format = FMT_CURRENCY; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
    r += 2

# Section C: DCF
r += 1
style_section_header(ws_pltr, r, 1, 10, "SECTION C — DCF VALUATION", DARK_NAVY)
r += 1

ws_pltr.cell(r,1).value = "Cash + US Treasuries ($M)"
ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
ws_pltr.cell(r,2).value = 8027; ws_pltr.cell(r,2).font = font(BLUE)
ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
add_comment(ws_pltr.cell(r,2), "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026")
pltr_cash_row = r; r += 1

ws_pltr.cell(r,1).value = "Total Debt ($M)"
ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
ws_pltr.cell(r,2).value = 0; ws_pltr.cell(r,2).font = font(BLUE)
ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
add_comment(ws_pltr.cell(r,2), "Source: Palantir Q1 2026 Earnings Press Release — No debt outstanding")
pltr_debt_row = r; r += 1

ws_pltr.cell(r,1).value = "Diluted Shares Outstanding (M)"
ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
ws_pltr.cell(r,2).value = 2571; ws_pltr.cell(r,2).font = font(BLUE)
ws_pltr.cell(r,2).number_format = FMT_NUM; ws_pltr.cell(r,2).alignment = center()
add_comment(ws_pltr.cell(r,2), "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026")
pltr_shares_row = r; r += 1

ws_pltr.cell(r,1).value = "Current Share Price ($)"
ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
ws_pltr.cell(r,2).value = 141.51; ws_pltr.cell(r,2).font = font(BLUE)
ws_pltr.cell(r,2).number_format = FMT_PRICE; ws_pltr.cell(r,2).alignment = center()
add_comment(ws_pltr.cell(r,2), "Source: Market price as of June 2026 analysis date")
pltr_price_row = r; r += 2

dcf_inputs_pltr = {
    "Bull": (0.12, 0.04, 0.55),
    "Base": (0.13, 0.035, 0.50),
    "Bear": (0.15, 0.03, 0.40),
}
pltr_dcf_rows = {}

for scen in scenarios:
    disc, tgr, tfcf = dcf_inputs_pltr[scen]
    style_section_header(ws_pltr, r, 1, 10, f"{scen} DCF — Assumptions & Output", scenario_colors[scen])
    r += 1

    ws_pltr.cell(r,1).value = "Discount Rate (WACC)"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = disc; ws_pltr.cell(r,2).font = font(BLUE)
    ws_pltr.cell(r,2).number_format = FMT_PCT; ws_pltr.cell(r,2).alignment = center()
    add_comment(ws_pltr.cell(r,2), f"Source: Analyst assumption; {scen} WACC for PLTR")
    disc_row = r; r += 1

    ws_pltr.cell(r,1).value = "Terminal Growth Rate"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = tgr; ws_pltr.cell(r,2).font = font(BLUE)
    ws_pltr.cell(r,2).number_format = FMT_PCT; ws_pltr.cell(r,2).alignment = center()
    add_comment(ws_pltr.cell(r,2), f"Source: Analyst assumption; {scen} terminal growth rate")
    tgr_row = r; r += 1

    ws_pltr.cell(r,1).value = "Terminal FCF Margin"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = tfcf; ws_pltr.cell(r,2).font = font(BLUE)
    ws_pltr.cell(r,2).number_format = FMT_PCT; ws_pltr.cell(r,2).alignment = center()
    add_comment(ws_pltr.cell(r,2), f"Source: Analyst assumption; {scen} normalized FCF margin")
    tfcf_row = r; r += 1

    d_ref = f"$B${disc_row}"
    g_ref = f"$B${tgr_row}"
    m_ref = f"$B${tfcf_row}"
    fcf_row = pltr_fcf_rows[scen]
    rev_row = pltr_rev_rows[scen]

    ws_pltr.cell(r,1).value = "PV of FCFs ($M) — Year 1 to 5"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    for yi, col in enumerate(["B","C","D","E","F"], 1):
        cell = ws_pltr.cell(r, yi+1)
        cell.value = f"={col}{fcf_row}/((1+{d_ref})^{yi})"
        cell.font = font(BLACK); cell.number_format = FMT_CURRENCY; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
    pv_fcf_row = r; r += 1

    ws_pltr.cell(r,1).value = "Sum of PV FCFs ($M)"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    refs = ",".join([f"{get_column_letter(i+2)}{pv_fcf_row}" for i in range(5)])
    ws_pltr.cell(r,2).value = f"=SUM({refs})"
    ws_pltr.cell(r,2).font = font(BLACK); ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
    sum_pv_row = r; r += 1

    ws_pltr.cell(r,1).value = "Terminal Year FCF ($M)"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = f"=F{rev_row}*{m_ref}"
    ws_pltr.cell(r,2).font = font(BLACK); ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
    term_fcf_row = r; r += 1

    ws_pltr.cell(r,1).value = "Terminal Value ($M) [Gordon Growth]"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = f"=B{term_fcf_row}*(1+{g_ref})/({d_ref}-{g_ref})"
    ws_pltr.cell(r,2).font = font(BLACK); ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
    tv_row = r; r += 1

    ws_pltr.cell(r,1).value = "PV of Terminal Value ($M)"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = f"=B{tv_row}/((1+{d_ref})^5)"
    ws_pltr.cell(r,2).font = font(BLACK); ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
    pv_tv_row = r; r += 1

    ws_pltr.cell(r,1).value = "Enterprise Value ($M)"
    ws_pltr.cell(r,1).font = font(BLACK, bold=True); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = f"=B{sum_pv_row}+B{pv_tv_row}"
    ws_pltr.cell(r,2).font = font(BLACK, bold=True)
    ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
    ws_pltr.cell(r,2).fill = fill(LIGHT_GOLD)
    ev_row = r; r += 1

    ws_pltr.cell(r,1).value = "Equity Value ($M) [EV + Cash - Debt]"
    ws_pltr.cell(r,1).font = font(BLACK, bold=True); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = f"=B{ev_row}+$B${pltr_cash_row}-$B${pltr_debt_row}"
    ws_pltr.cell(r,2).font = font(BLACK, bold=True)
    ws_pltr.cell(r,2).number_format = FMT_CURRENCY; ws_pltr.cell(r,2).alignment = center()
    ws_pltr.cell(r,2).fill = fill(LIGHT_GOLD)
    eq_row = r; r += 1

    ws_pltr.cell(r,1).value = "Intrinsic Value Per Share ($)"
    ws_pltr.cell(r,1).font = font(BLACK, bold=True); ws_pltr.cell(r,1).alignment = left()
    # equity value in $M, shares in M -> price = (EqVal_M / Shares_M) * 1 (both in millions, ratio gives $)
    ws_pltr.cell(r,2).value = f"=B{eq_row}/$B${pltr_shares_row}"
    ws_pltr.cell(r,2).font = font(BLACK, bold=True)
    ws_pltr.cell(r,2).number_format = FMT_PRICE; ws_pltr.cell(r,2).alignment = center()
    ws_pltr.cell(r,2).fill = fill(GOLD)
    pltr_dcf_rows[scen] = r; r += 1

    ws_pltr.cell(r,1).value = "% Premium(-)/Discount(+) vs Current Price"
    ws_pltr.cell(r,1).font = font(BLACK); ws_pltr.cell(r,1).alignment = left()
    ws_pltr.cell(r,2).value = f"=($B${pltr_price_row}-B{pltr_dcf_rows[scen]})/B{pltr_dcf_rows[scen]}"
    ws_pltr.cell(r,2).font = font(BLACK); ws_pltr.cell(r,2).number_format = FMT_PCT; ws_pltr.cell(r,2).alignment = center()
    r += 2

# Section D: Rule of 40
style_section_header(ws_pltr, r, 1, 10, "SECTION D — RULE OF 40 (HISTORICAL)", DARK_NAVY)
r += 1
for ci, h in enumerate(["Period","Revenue Growth","Adj. Op Margin","Rule of 40 Score"], 1):
    ws_pltr.cell(r,ci).value = h
style_header_row(ws_pltr, r, 1, 5)
r += 1

ro40_pltr = [
    ("FY2025",   0.56, 0.50, "Source: Palantir FY2025 Annual Report"),
    ("Q1 2026",  0.85, 0.60, "Source: Palantir Q1 2026 Earnings Press Release, May 4 2026"),
]
for period, rg, om, src in ro40_pltr:
    ws_pltr.cell(r,1).value = period; ws_pltr.cell(r,1).font = font(BLUE)
    ws_pltr.cell(r,1).alignment = center()
    add_comment(ws_pltr.cell(r,1), src)
    ws_pltr.cell(r,2).value = rg; ws_pltr.cell(r,2).font = font(BLUE)
    ws_pltr.cell(r,2).number_format = FMT_PCT; ws_pltr.cell(r,2).alignment = center()
    add_comment(ws_pltr.cell(r,2), src)
    ws_pltr.cell(r,3).value = om; ws_pltr.cell(r,3).font = font(BLUE)
    ws_pltr.cell(r,3).number_format = FMT_PCT; ws_pltr.cell(r,3).alignment = center()
    add_comment(ws_pltr.cell(r,3), src)
    ws_pltr.cell(r,4).value = f"=(B{r}+C{r})*100"
    ws_pltr.cell(r,4).font = font(BLACK, bold=True)
    ws_pltr.cell(r,4).number_format = FMT_RO40; ws_pltr.cell(r,4).alignment = center()
    ws_pltr.cell(r,4).fill = fill(LIGHT_GOLD)
    r += 1

# ════════════════════════════════════════════════════════
# SHEET 3: NOW_Model
# ════════════════════════════════════════════════════════
ws_now = wb.create_sheet("NOW_Model")
set_col_widths(ws_now, {"A":36,"B":14,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14})
freeze(ws_now)

r = 1
ws_now.merge_cells(f"A{r}:J{r}")
c = ws_now.cell(r,1,"SERVICENOW (NOW) — Investment Analysis Model")
c.font = font(WHITE, bold=True, size=13); c.fill = fill(TEAL); c.alignment = center()
ws_now.row_dimensions[r].height = 28

r += 1
style_section_header(ws_now, r, 1, 10, "SECTION A — HISTORICAL DATA (FY2025–Q1 2026)", TEAL)
r += 1

for ci, h in enumerate(["Metric","FY2025","Q1 2026"], 1):
    ws_now.cell(r,ci).value = h
style_header_row(ws_now, r, 1, 10)
r += 1

now_hist = [
    ("Subscription Revenue ($M)",   [12840, 3671],  FMT_CURRENCY,
     "Source: ServiceNow FY2025 Annual Report; Q1 2026 Earnings Press Release"),
    ("Total Revenue ($M)",          [13285, 3770],  FMT_CURRENCY,
     "Source: ServiceNow FY2025 Annual Report (est.); Q1 2026 Press Release"),
    ("Sub Revenue YoY Growth",      [0.21, 0.22],   FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("GAAP Gross Margin",           [None, 0.75],   FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Non-GAAP Gross Margin",       [None, 0.795],  FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Sub Gross Margin (GAAP)",     [None, 0.775],  FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Sub Gross Margin (Non-GAAP)", [None, 0.815],  FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Non-GAAP Operating Margin",   [None, 0.32],   FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Non-GAAP FCF Margin",         [0.40, 0.44],   FMT_PCT,
     "Source: ServiceNow FY2025 Annual Report (est.); Q1 2026 Press Release"),
    ("cRPO ($M)",                   [None, 12640],  FMT_CURRENCY,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("cRPO YoY Growth",             [None, 0.225],  FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Total RPO ($M)",              [None, 27700],  FMT_CURRENCY,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Total RPO YoY Growth",        [None, 0.25],   FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("SBC ($M)",                    [None, 547],    FMT_CURRENCY,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("SBC % of Revenue",            [None, 0.145],  FMT_PCT,
     "Source: ServiceNow Q1 2026 Earnings Press Release"),
]
for label, vals, fmt, src in now_hist:
    ws_now.cell(r,1).value = label; ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    if r % 2 == 0: ws_now.cell(r,1).fill = fill(LIGHT_GREY)
    for ci, v in enumerate(vals, 2):
        cell = ws_now.cell(r,ci)
        if v is not None:
            cell.value = v; cell.font = font(BLUE)
            add_comment(cell, src)
        cell.number_format = fmt; cell.alignment = center()
        if r % 2 == 0: cell.fill = fill(LIGHT_GREY)
    r += 1

# Section D: Key SaaS Metrics
r += 1
style_section_header(ws_now, r, 1, 10, "SECTION D — KEY SaaS METRICS (Q1 2026)", TEAL)
r += 1
for ci, h in enumerate(["Metric","Value","Growth / Note"], 1):
    ws_now.cell(r,ci).value = h
style_header_row(ws_now, r, 1, 10)
r += 1

saas_metrics = [
    ("Customers >$5M ACV",             630,  "+22% YoY",  "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Transactions >$5M ACV (Q1)",      16,   "+80% YoY",  "Source: ServiceNow Q1 2026 Earnings Press Release"),
    ("Now Assist >$1M ACV Customers",   None, "+130% YoY", "Source: ServiceNow Q1 2026 Earnings Press Release"),
]
for label, val, note, src in saas_metrics:
    ws_now.cell(r,1).value = label; ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    if val is not None:
        ws_now.cell(r,2).value = val; ws_now.cell(r,2).font = font(BLUE)
        ws_now.cell(r,2).number_format = FMT_NUM; ws_now.cell(r,2).alignment = center()
        add_comment(ws_now.cell(r,2), src)
    ws_now.cell(r,3).value = note; ws_now.cell(r,3).font = font(BLUE); ws_now.cell(r,3).alignment = center()
    add_comment(ws_now.cell(r,3), src)
    if r % 2 == 0:
        for c in range(1,11): ws_now.cell(r,c).fill = fill(LIGHT_GREY)
    r += 1

# Section B: NOW Projections
r += 2
style_section_header(ws_now, r, 1, 10, "SECTION B — REVENUE & FCF PROJECTIONS (FY2026–FY2030)", DARK_NAVY)
r += 1

ws_now.cell(r,1).value = "FY2025 Total Revenue ($M) [Base year]"
ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
ws_now.cell(r,2).value = 13285; ws_now.cell(r,2).font = font(BLUE)
ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
add_comment(ws_now.cell(r,2), "Source: ServiceNow FY2025 Annual Report (estimate)")
now_fy25_rev_row = r; r += 1

ws_now.cell(r,1).value = "FY2026 Guided Revenue ($M) [Sub $15,755M + PS ~$345M]"
ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
ws_now.cell(r,2).value = 16100; ws_now.cell(r,2).font = font(BLUE)
ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
add_comment(ws_now.cell(r,2), "Source: ServiceNow FY2026 guidance; subscription revenue $15,755M + PS ~$345M")
now_fy26_rev_row = r; r += 1

ws_now.cell(r,1).value = "Year"
for ci, yr in enumerate(years, 2): ws_now.cell(r,ci).value = yr
style_header_row(ws_now, r, 1, 7)
r += 1

now_growth_rates = {
    "Bull": [0.22, 0.22, 0.20, 0.18, 0.16],
    "Base": [0.22, 0.20, 0.18, 0.15, 0.12],
    "Bear": [0.22, 0.15, 0.12, 0.10, 0.08],
}
now_fcf_margins = {
    "Bull": [0.37, 0.40, 0.42, 0.43, 0.44],
    "Base": [0.35, 0.37, 0.39, 0.40, 0.41],
    "Bear": [0.32, 0.33, 0.34, 0.35, 0.36],
}
now_growth_rows = {}; now_fcf_margin_rows = {}; now_rev_rows = {}; now_fcf_rows = {}

for scen in scenarios:
    style_section_header(ws_now, r, 1, 10, f"{scen} Scenario", scenario_colors[scen])
    r += 1

    ws_now.cell(r,1).value = f"{scen} Rev Growth Assumption"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    now_growth_rows[scen] = r
    src_g = f"Source: ServiceNow FY2026 guidance ~22%; {scen} scenario analyst estimates"
    for ci, g in enumerate(now_growth_rates[scen], 2):
        cell = ws_now.cell(r,ci)
        cell.value = g; cell.font = font(BLUE)
        cell.number_format = FMT_PCT; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
        add_comment(cell, src_g)
    r += 1

    ws_now.cell(r,1).value = f"{scen} Revenue ($M)"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    now_rev_rows[scen] = r
    for ci in range(2, 7):
        col = get_column_letter(ci)
        cell = ws_now.cell(r,ci)
        if ci == 2:
            # FY2026 uses the guided revenue directly (growth already embedded in $16,100M)
            cell.value = f"=$B${now_fy26_rev_row}"
        else:
            prev = get_column_letter(ci-1)
            cell.value = f"={prev}{r}*(1+{col}{now_growth_rows[scen]})"
        cell.font = font(BLACK); cell.number_format = FMT_CURRENCY; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
    r += 1

    ws_now.cell(r,1).value = f"{scen} FCF Margin Assumption"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    now_fcf_margin_rows[scen] = r
    src_m = f"Source: NOW Q1 2026 non-GAAP FCF margin 44%; {scen} scenario analyst estimates"
    for ci, m in enumerate(now_fcf_margins[scen], 2):
        cell = ws_now.cell(r,ci)
        cell.value = m; cell.font = font(BLUE)
        cell.number_format = FMT_PCT; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
        add_comment(cell, src_m)
    r += 1

    ws_now.cell(r,1).value = f"{scen} FCF ($M)"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    now_fcf_rows[scen] = r
    for ci in range(2, 7):
        col = get_column_letter(ci)
        cell = ws_now.cell(r,ci)
        cell.value = f"={col}{now_rev_rows[scen]}*{col}{now_fcf_margin_rows[scen]}"
        cell.font = font(BLACK); cell.number_format = FMT_CURRENCY; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
    r += 2

# Section C: NOW DCF
r += 1
style_section_header(ws_now, r, 1, 10, "SECTION C — DCF VALUATION", DARK_NAVY)
r += 1

ws_now.cell(r,1).value = "Net Cash Position ($M) [Cash+Mkt Securities-Debt]"
ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
ws_now.cell(r,2).value = 3691; ws_now.cell(r,2).font = font(BLUE)
ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
add_comment(ws_now.cell(r,2), "Source: ServiceNow Q1 2026 Balance Sheet: Cash $2,702M + Mkt Securities $2,480M - Debt $1,491M = $3,691M")
now_cash_row = r; r += 1

ws_now.cell(r,1).value = "Diluted Shares Outstanding (M)"
ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
ws_now.cell(r,2).value = 1040; ws_now.cell(r,2).font = font(BLUE)
ws_now.cell(r,2).number_format = FMT_NUM; ws_now.cell(r,2).alignment = center()
add_comment(ws_now.cell(r,2), "Source: ServiceNow Q1 2026 Earnings Press Release")
now_shares_row = r; r += 1

ws_now.cell(r,1).value = "Current Share Price ($)"
ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
ws_now.cell(r,2).value = 119.29; ws_now.cell(r,2).font = font(BLUE)
ws_now.cell(r,2).number_format = FMT_PRICE; ws_now.cell(r,2).alignment = center()
add_comment(ws_now.cell(r,2), "Source: Market price as of June 2026 analysis date")
now_price_row = r; r += 2

dcf_inputs_now = {
    "Bull": (0.10, 0.04, 0.42),
    "Base": (0.11, 0.035, 0.40),
    "Bear": (0.12, 0.03, 0.35),
}
now_dcf_rows = {}

for scen in scenarios:
    disc, tgr, tfcf = dcf_inputs_now[scen]
    style_section_header(ws_now, r, 1, 10, f"{scen} DCF — Assumptions & Output", scenario_colors[scen])
    r += 1

    ws_now.cell(r,1).value = "Discount Rate (WACC)"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = disc; ws_now.cell(r,2).font = font(BLUE)
    ws_now.cell(r,2).number_format = FMT_PCT; ws_now.cell(r,2).alignment = center()
    add_comment(ws_now.cell(r,2), f"Source: Analyst assumption; {scen} WACC for NOW")
    disc_row = r; r += 1

    ws_now.cell(r,1).value = "Terminal Growth Rate"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = tgr; ws_now.cell(r,2).font = font(BLUE)
    ws_now.cell(r,2).number_format = FMT_PCT; ws_now.cell(r,2).alignment = center()
    add_comment(ws_now.cell(r,2), f"Source: Analyst assumption; {scen} terminal growth rate")
    tgr_row = r; r += 1

    ws_now.cell(r,1).value = "Terminal FCF Margin"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = tfcf; ws_now.cell(r,2).font = font(BLUE)
    ws_now.cell(r,2).number_format = FMT_PCT; ws_now.cell(r,2).alignment = center()
    add_comment(ws_now.cell(r,2), f"Source: Analyst assumption; {scen} normalized FCF margin")
    tfcf_row = r; r += 1

    d_ref = f"$B${disc_row}"
    g_ref = f"$B${tgr_row}"
    m_ref = f"$B${tfcf_row}"
    fcf_row = now_fcf_rows[scen]
    rev_row = now_rev_rows[scen]

    ws_now.cell(r,1).value = "PV of FCFs ($M) — Year 1 to 5"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    for yi, col in enumerate(["B","C","D","E","F"], 1):
        cell = ws_now.cell(r, yi+1)
        cell.value = f"={col}{fcf_row}/((1+{d_ref})^{yi})"
        cell.font = font(BLACK); cell.number_format = FMT_CURRENCY; cell.alignment = center()
        cell.fill = fill(scenario_light[scen])
    pv_fcf_row = r; r += 1

    refs = ",".join([f"{get_column_letter(i+2)}{pv_fcf_row}" for i in range(5)])
    ws_now.cell(r,1).value = "Sum of PV FCFs ($M)"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = f"=SUM({refs})"
    ws_now.cell(r,2).font = font(BLACK); ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
    sum_pv_row = r; r += 1

    ws_now.cell(r,1).value = "Terminal Year FCF ($M)"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = f"=F{rev_row}*{m_ref}"
    ws_now.cell(r,2).font = font(BLACK); ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
    term_fcf_row = r; r += 1

    ws_now.cell(r,1).value = "Terminal Value ($M) [Gordon Growth]"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = f"=B{term_fcf_row}*(1+{g_ref})/({d_ref}-{g_ref})"
    ws_now.cell(r,2).font = font(BLACK); ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
    tv_row = r; r += 1

    ws_now.cell(r,1).value = "PV of Terminal Value ($M)"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = f"=B{tv_row}/((1+{d_ref})^5)"
    ws_now.cell(r,2).font = font(BLACK); ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
    pv_tv_row = r; r += 1

    ws_now.cell(r,1).value = "Enterprise Value ($M)"
    ws_now.cell(r,1).font = font(BLACK, bold=True); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = f"=B{sum_pv_row}+B{pv_tv_row}"
    ws_now.cell(r,2).font = font(BLACK, bold=True)
    ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
    ws_now.cell(r,2).fill = fill(LIGHT_GOLD)
    ev_row = r; r += 1

    ws_now.cell(r,1).value = "Equity Value ($M) [EV + Net Cash]"
    ws_now.cell(r,1).font = font(BLACK, bold=True); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = f"=B{ev_row}+$B${now_cash_row}"
    ws_now.cell(r,2).font = font(BLACK, bold=True)
    ws_now.cell(r,2).number_format = FMT_CURRENCY; ws_now.cell(r,2).alignment = center()
    ws_now.cell(r,2).fill = fill(LIGHT_GOLD)
    eq_row = r; r += 1

    ws_now.cell(r,1).value = "Intrinsic Value Per Share ($)"
    ws_now.cell(r,1).font = font(BLACK, bold=True); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = f"=B{eq_row}/$B${now_shares_row}"
    ws_now.cell(r,2).font = font(BLACK, bold=True)
    ws_now.cell(r,2).number_format = FMT_PRICE; ws_now.cell(r,2).alignment = center()
    ws_now.cell(r,2).fill = fill(GOLD)
    now_dcf_rows[scen] = r; r += 1

    ws_now.cell(r,1).value = "% Premium(-)/Discount(+) vs Current Price"
    ws_now.cell(r,1).font = font(BLACK); ws_now.cell(r,1).alignment = left()
    ws_now.cell(r,2).value = f"=($B${now_price_row}-B{now_dcf_rows[scen]})/B{now_dcf_rows[scen]}"
    ws_now.cell(r,2).font = font(BLACK); ws_now.cell(r,2).number_format = FMT_PCT; ws_now.cell(r,2).alignment = center()
    r += 2

# ════════════════════════════════════════════════════════
# SHEET 4: Comparison
# ════════════════════════════════════════════════════════
ws_comp = wb.create_sheet("Comparison")
set_col_widths(ws_comp, {"A":22,"B":16,"C":16,"D":16,"E":16,"F":16,"G":32})
freeze(ws_comp)

r = 1
ws_comp.merge_cells(f"A{r}:G{r}")
c = ws_comp.cell(r,1,"PLTR vs NOW — Comparative Analysis")
c.font = font(WHITE, bold=True, size=13); c.fill = fill(DARK_NAVY); c.alignment = center()
ws_comp.row_dimensions[r].height = 28

# Rule of 40 table
r += 2
style_section_header(ws_comp, r, 1, 7, "RULE OF 40 COMPARISON (Q1 2026)", DARK_NAVY)
r += 1
for ci, h in enumerate(["Company","Revenue Growth","FCF/Op Margin","Rule of 40","Forward EV/S","EV/FCF","Notes"], 1):
    ws_comp.cell(r,ci).value = h
style_header_row(ws_comp, r, 1, 7)
r += 1

ro40_src = "Source: Palantir/ServiceNow Q1 2026 Press Releases; EV/S and EV/FCF are analyst estimates"
ro40_comp_data = [
    ("PLTR (Q1 2026)", 0.85, 0.60, 46.5, 82.7, "Adj. op margin; AI infrastructure"),
    ("NOW (Q1 2026)",  0.22, 0.32, 7.5,  21.3, "Non-GAAP op margin; Enterprise workflow"),
]
for co, rg, om, ev_s, ev_fcf, note in ro40_comp_data:
    ws_comp.cell(r,1).value = co; ws_comp.cell(r,1).font = font(BLUE); ws_comp.cell(r,1).alignment = left()
    ws_comp.cell(r,2).value = rg; ws_comp.cell(r,2).font = font(BLUE)
    ws_comp.cell(r,2).number_format = FMT_PCT; ws_comp.cell(r,2).alignment = center()
    add_comment(ws_comp.cell(r,2), ro40_src)
    ws_comp.cell(r,3).value = om; ws_comp.cell(r,3).font = font(BLUE)
    ws_comp.cell(r,3).number_format = FMT_PCT; ws_comp.cell(r,3).alignment = center()
    add_comment(ws_comp.cell(r,3), ro40_src)
    ws_comp.cell(r,4).value = f"=(B{r}+C{r})*100"
    ws_comp.cell(r,4).font = font(BLACK, bold=True)
    ws_comp.cell(r,4).number_format = FMT_RO40; ws_comp.cell(r,4).alignment = center()
    ws_comp.cell(r,4).fill = fill(LIGHT_GOLD)
    ws_comp.cell(r,5).value = ev_s; ws_comp.cell(r,5).font = font(BLUE)
    ws_comp.cell(r,5).number_format = FMT_MULT; ws_comp.cell(r,5).alignment = center()
    add_comment(ws_comp.cell(r,5), ro40_src)
    ws_comp.cell(r,6).value = ev_fcf; ws_comp.cell(r,6).font = font(BLUE)
    ws_comp.cell(r,6).number_format = FMT_MULT; ws_comp.cell(r,6).alignment = center()
    add_comment(ws_comp.cell(r,6), ro40_src)
    ws_comp.cell(r,7).value = note; ws_comp.cell(r,7).font = font(BLACK, italic=True); ws_comp.cell(r,7).alignment = left()
    if r % 2 == 0:
        for c in range(1,8): ws_comp.cell(r,c).fill = fill(LIGHT_GREY)
    r += 1

# Peer comps
r += 2
style_section_header(ws_comp, r, 1, 7, "PEER COMPARABLES TABLE (Approximate, June 2026 — VERIFY BEFORE USE)", DARK_NAVY)
r += 1
for ci, h in enumerate(["Company","Ticker","Fwd Rev Growth","FCF Margin","Rule of 40","Fwd EV/S","Segment / Comment"], 1):
    ws_comp.cell(r,ci).value = h
style_header_row(ws_comp, r, 1, 7)
r += 1

peer_src = "Source: Analyst/consensus estimates approx. June 2026. Verify against Bloomberg/FactSet before use."
peers = [
    ("Palantir",    "PLTR", 0.71, 0.56, 46.5, "AI infrastructure / data analytics"),
    ("ServiceNow",  "NOW",  0.22, 0.35, 7.5,  "Enterprise workflow / AI platform"),
    ("Salesforce",  "CRM",  0.09, 0.32, 6.5,  "CRM / cloud applications"),
    ("Snowflake",   "SNOW", 0.23, 0.18, 10.0, "Data cloud / analytics platform"),
    ("Datadog",     "DDOG", 0.23, 0.25, 14.0, "Observability / monitoring"),
    ("CrowdStrike", "CRWD", 0.20, 0.28, 20.0, "Cybersecurity / endpoint protection"),
    ("Workday",     "WDAY", 0.14, 0.25, 6.0,  "HCM / Finance cloud"),
]
for co, tkr, rg, fcf_m, ev_s, comment in peers:
    ws_comp.cell(r,1).value = co; ws_comp.cell(r,1).font = font(BLUE); ws_comp.cell(r,1).alignment = left()
    ws_comp.cell(r,2).value = tkr; ws_comp.cell(r,2).font = font(BLUE); ws_comp.cell(r,2).alignment = center()
    ws_comp.cell(r,3).value = rg; ws_comp.cell(r,3).font = font(BLUE)
    ws_comp.cell(r,3).number_format = FMT_PCT; ws_comp.cell(r,3).alignment = center()
    add_comment(ws_comp.cell(r,3), peer_src)
    ws_comp.cell(r,4).value = fcf_m; ws_comp.cell(r,4).font = font(BLUE)
    ws_comp.cell(r,4).number_format = FMT_PCT; ws_comp.cell(r,4).alignment = center()
    add_comment(ws_comp.cell(r,4), peer_src)
    ws_comp.cell(r,5).value = f"=(C{r}+D{r})*100"
    ws_comp.cell(r,5).font = font(BLACK, bold=True)
    ws_comp.cell(r,5).number_format = FMT_RO40; ws_comp.cell(r,5).alignment = center()
    ws_comp.cell(r,5).fill = fill(LIGHT_GOLD)
    ws_comp.cell(r,6).value = ev_s; ws_comp.cell(r,6).font = font(BLUE)
    ws_comp.cell(r,6).number_format = FMT_MULT; ws_comp.cell(r,6).alignment = center()
    add_comment(ws_comp.cell(r,6), peer_src)
    ws_comp.cell(r,7).value = comment; ws_comp.cell(r,7).font = font(BLACK, italic=True); ws_comp.cell(r,7).alignment = left()
    if r % 2 == 0:
        for c in range(1,8): ws_comp.cell(r,c).fill = fill(LIGHT_GREY)
    r += 1

r += 1
ws_comp.merge_cells(f"A{r}:G{r}")
ws_comp.cell(r,1).value = "NOTE: Peer comps are approximate estimates as of June 2026. Verify against current Bloomberg/FactSet before use."
ws_comp.cell(r,1).font = font(BLACK, italic=True, size=9); ws_comp.cell(r,1).alignment = left()

# Valuation summary
r += 2
style_section_header(ws_comp, r, 1, 7, "VALUATION SUMMARY — CURRENT PRICE vs DCF INTRINSIC VALUE", DARK_NAVY)
r += 1
for ci, h in enumerate(["Company","Current Price","DCF Bull","DCF Base","DCF Bear","vs Bull (%)","vs Base (%)"], 1):
    ws_comp.cell(r,ci).value = h
style_header_row(ws_comp, r, 1, 7)
r += 1

# PLTR row
ws_comp.cell(r,1).value = "PLTR"; ws_comp.cell(r,1).font = font(BLACK, bold=True); ws_comp.cell(r,1).alignment = left()
ws_comp.cell(r,2).value = f"=PLTR_Model!$B${pltr_price_row}"; ws_comp.cell(r,2).font = font(GREEN)
ws_comp.cell(r,2).number_format = FMT_PRICE; ws_comp.cell(r,2).alignment = center()
ws_comp.cell(r,3).value = f"=PLTR_Model!$B${pltr_dcf_rows['Bull']}"; ws_comp.cell(r,3).font = font(GREEN)
ws_comp.cell(r,3).number_format = FMT_PRICE; ws_comp.cell(r,3).alignment = center()
ws_comp.cell(r,4).value = f"=PLTR_Model!$B${pltr_dcf_rows['Base']}"; ws_comp.cell(r,4).font = font(GREEN)
ws_comp.cell(r,4).number_format = FMT_PRICE; ws_comp.cell(r,4).alignment = center()
ws_comp.cell(r,5).value = f"=PLTR_Model!$B${pltr_dcf_rows['Bear']}"; ws_comp.cell(r,5).font = font(GREEN)
ws_comp.cell(r,5).number_format = FMT_PRICE; ws_comp.cell(r,5).alignment = center()
ws_comp.cell(r,6).value = f"=(B{r}-C{r})/C{r}"; ws_comp.cell(r,6).font = font(BLACK)
ws_comp.cell(r,6).number_format = FMT_PCT; ws_comp.cell(r,6).alignment = center()
ws_comp.cell(r,7).value = f"=(B{r}-D{r})/D{r}"; ws_comp.cell(r,7).font = font(BLACK)
ws_comp.cell(r,7).number_format = FMT_PCT; ws_comp.cell(r,7).alignment = center()
r += 1

# NOW row
ws_comp.cell(r,1).value = "NOW"; ws_comp.cell(r,1).font = font(BLACK, bold=True); ws_comp.cell(r,1).alignment = left()
ws_comp.cell(r,2).value = f"=NOW_Model!$B${now_price_row}"; ws_comp.cell(r,2).font = font(GREEN)
ws_comp.cell(r,2).number_format = FMT_PRICE; ws_comp.cell(r,2).alignment = center()
ws_comp.cell(r,3).value = f"=NOW_Model!$B${now_dcf_rows['Bull']}"; ws_comp.cell(r,3).font = font(GREEN)
ws_comp.cell(r,3).number_format = FMT_PRICE; ws_comp.cell(r,3).alignment = center()
ws_comp.cell(r,4).value = f"=NOW_Model!$B${now_dcf_rows['Base']}"; ws_comp.cell(r,4).font = font(GREEN)
ws_comp.cell(r,4).number_format = FMT_PRICE; ws_comp.cell(r,4).alignment = center()
ws_comp.cell(r,5).value = f"=NOW_Model!$B${now_dcf_rows['Bear']}"; ws_comp.cell(r,5).font = font(GREEN)
ws_comp.cell(r,5).number_format = FMT_PRICE; ws_comp.cell(r,5).alignment = center()
ws_comp.cell(r,6).value = f"=(B{r}-C{r})/C{r}"; ws_comp.cell(r,6).font = font(BLACK)
ws_comp.cell(r,6).number_format = FMT_PCT; ws_comp.cell(r,6).alignment = center()
ws_comp.cell(r,7).value = f"=(B{r}-D{r})/D{r}"; ws_comp.cell(r,7).font = font(BLACK)
ws_comp.cell(r,7).number_format = FMT_PCT; ws_comp.cell(r,7).alignment = center()

# ════════════════════════════════════════════════════════
# SHEET 5: Options_Strategy
# ════════════════════════════════════════════════════════
ws_opt = wb.create_sheet("Options_Strategy")
set_col_widths(ws_opt, {"A":24,"B":8,"C":22,"D":14,"E":14,"F":14,"G":14,"H":14,"I":38})
freeze(ws_opt)

r = 1
ws_opt.merge_cells(f"A{r}:I{r}")
c = ws_opt.cell(r,1,"OPTIONS STRATEGY — PLTR & NOW (June 2026)")
c.font = font(WHITE, bold=True, size=13); c.fill = fill(DARK_NAVY); c.alignment = center()
ws_opt.row_dimensions[r].height = 28

r += 2
style_section_header(ws_opt, r, 1, 9, "MARKET DATA REFERENCE (June 2026)", MID_BLUE)
r += 1
for ci, h in enumerate(["Ticker","Price","IV (%)","IV Rank","200-day SMA","52-wk Low","52-wk High","RSI (est.)","Notes"], 1):
    ws_opt.cell(r,ci).value = h
style_header_row(ws_opt, r, 1, 9)
r += 1

mkt_src = "Source: Market data as of June 2026 analysis date; verify with live data before trading"
pltr_mkt_row = [("PLTR", 141.51, 0.6879, 46.73, 162.0, None,  None,   None, "IV 68.79%; IVR 46.73; trades ~12% below 200-day SMA of $162")]
now_mkt_row  = [("NOW",  119.29, None,   None,  None,  81.24, 211.48, None, "52-wk range $81.24-$211.48; near lower historical range")]

for tkr, price, iv, ivr, sma, lo, hi, rsi, note in pltr_mkt_row + now_mkt_row:
    vals = [tkr, price, iv, ivr, sma, lo, hi, rsi, note]
    fmts = [FMT_TEXT, FMT_PRICE, FMT_PCT, "#,##0.0", FMT_PRICE, FMT_PRICE, FMT_PRICE, "#,##0.0", FMT_TEXT]
    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws_opt.cell(r,ci)
        if v is not None:
            cell.value = v; cell.font = font(BLUE)
            add_comment(cell, mkt_src)
        cell.number_format = fmt; cell.alignment = center() if ci > 1 else left()
    r += 1

r += 2
style_section_header(ws_opt, r, 1, 9, "OPTIONS STRATEGIES SUMMARY TABLE", DARK_NAVY)
r += 1
opt_hdrs = ["Strategy","Ticker","Structure","Strike(s)","Expiry","Premium (est.)","Max Gain","Max Loss","Rationale"]
for ci, h in enumerate(opt_hdrs, 1):
    ws_opt.cell(r,ci).value = h
style_header_row(ws_opt, r, 1, 9)
r += 1

opt_src = "Source: Analyst options strategy estimates as of June 2026. Verify with live quotes before execution."
strategies = [
    ("1 — Bear Call Spread",  "PLTR", "Sell $155 Call / Buy $165 Call", "$155 / $165", "Aug 2026 (~2mo)", "~$3-4 credit",   "~$3-4 credit",  "~$6-7 per spread", "PLTR trades far above DCF intrinsic; moderate IVR 46.73 supports premium selling; protects existing holders from stall scenarios"),
    ("2 — Covered Call",      "PLTR", "Sell OTM Call vs long shares",   "$155-160",    "6-8 weeks out",   "~$5-7/contract", "~$5-7 premium", "Uncapped above strike", "At ~69% IV, significant premium income; reduces cost basis; best for long holders not expecting near-term rally"),
    ("3 — Cash-Secured Put",  "PLTR", "Sell OTM Put / Hold cash",       "$120-125",    "~8 weeks out",    "~$8-10",         "~$8-10 premium","Assigned at ~$110-117", "Effective cost basis ~$110-117 if assigned (near DCF bull ~$78 zone); income if PLTR holds above strike"),
    ("1 — Long Call Spread",  "NOW",  "Buy $120 Call / Sell $150 Call", "$120 / $150", "Jan 2027 (~7mo)", "~$10-12 debit",  "~$18-20 at $150","~$10-12 debit",   "Defined risk bullish play; NOW at/below DCF base case; captures recovery toward $129+ intrinsic value; ~1.7:1 R/R"),
    ("2 — Cash-Secured Put",  "NOW",  "Sell OTM Put / Hold cash",       "$110",        "8-10 weeks out",  "~$7-9",          "~$7-9 premium", "Assigned at ~$101-103", "Entry well below DCF base $129; significant margin of safety if assigned; income generation otherwise"),
    ("3 — Covered Call Enh.", "NOW",  "Sell OTM Call vs long shares",   "$135-140",    "6-8 weeks out",   "~$4-6/contract", "~$4-6 premium", "Uncapped above strike", "Income enhancement for existing NOW holders; generates yield while holding quality name near lows"),
]

for strat in strategies:
    strategy, ticker, structure, strikes, expiry, premium, max_gain, max_loss, rationale = strat
    vals = [strategy, ticker, structure, strikes, expiry, premium, max_gain, max_loss, rationale]
    fill_color = LIGHT_BLUE if ticker == "PLTR" else LIGHT_TEAL
    for ci, v in enumerate(vals, 1):
        cell = ws_opt.cell(r,ci)
        cell.value = v; cell.font = font(BLUE); cell.number_format = FMT_TEXT
        cell.alignment = left() if ci in [1,3,9] else center()
        cell.fill = fill(fill_color)
        add_comment(cell, opt_src)
    r += 1

r += 2
ws_opt.merge_cells(f"A{r}:I{r}")
ws_opt.cell(r,1).value = "DISCLAIMER: Options strategies are for educational/analytical purposes only. Verify all premiums with live market data. Options trading involves significant risk of loss. Consult a licensed financial advisor before trading."
ws_opt.cell(r,1).font = font(BLACK, italic=True, size=9); ws_opt.cell(r,1).alignment = left()

# ════════════════════════════════════════════════════════
# SHEET 1: Dashboard (inserted first)
# ════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Dashboard", 0)
set_col_widths(ws_dash, {"A":30,"B":18,"C":20,"D":4,"E":30,"F":18,"G":20})
freeze(ws_dash)

r = 1
ws_dash.merge_cells(f"A{r}:G{r}")
c = ws_dash.cell(r,1,"INVESTMENT ANALYSIS DASHBOARD — PLTR vs NOW  |  June 2026")
c.font = font(WHITE, bold=True, size=14); c.fill = fill(DARK_NAVY); c.alignment = center()
ws_dash.row_dimensions[r].height = 34

r += 2
ws_dash.merge_cells(f"A{r}:C{r}")
ph = ws_dash.cell(r,1,"PALANTIR TECHNOLOGIES (PLTR)")
ph.font = font(WHITE, bold=True, size=12); ph.fill = fill(MID_BLUE); ph.alignment = center()
ws_dash.merge_cells(f"E{r}:G{r}")
nh = ws_dash.cell(r,5,"SERVICENOW (NOW)")
nh.font = font(WHITE, bold=True, size=12); nh.fill = fill(TEAL); nh.alignment = center()
ws_dash.row_dimensions[r].height = 20
r += 1

for ci, h in enumerate(["Metric","Value","Notes/Source"], 1):
    for base in [1, 5]:
        ws_dash.cell(r,base+ci-1).value = h
        ws_dash.cell(r,base+ci-1).font = font(WHITE, bold=True)
        ws_dash.cell(r,base+ci-1).fill = fill(GREY_HDR)
        ws_dash.cell(r,base+ci-1).alignment = center()
r += 1

def dash_row(ws, r, p_label, p_val, p_note, p_fmt, n_label, n_val, n_note, n_fmt, alt=False):
    # PLTR side
    ws.cell(r,1).value = p_label
    ws.cell(r,1).font = font(BLACK)
    ws.cell(r,1).alignment = left()

    cell_p = ws.cell(r,2)
    if isinstance(p_val, str) and p_val.startswith("="):
        cell_p.value = p_val; cell_p.font = font(GREEN)
    elif p_val is not None:
        cell_p.value = p_val; cell_p.font = font(BLUE)
    cell_p.number_format = p_fmt; cell_p.alignment = center()

    if p_note:
        ws.cell(r,3).value = p_note; ws.cell(r,3).font = font(BLACK, italic=True, size=9)
        ws.cell(r,3).alignment = left()

    ws.cell(r,4).fill = fill(MED_GREY)

    # NOW side
    ws.cell(r,5).value = n_label
    ws.cell(r,5).font = font(BLACK)
    ws.cell(r,5).alignment = left()

    cell_n = ws.cell(r,6)
    if isinstance(n_val, str) and n_val.startswith("="):
        cell_n.value = n_val; cell_n.font = font(GREEN)
    elif n_val is not None:
        cell_n.value = n_val; cell_n.font = font(BLUE)
    cell_n.number_format = n_fmt; cell_n.alignment = center()

    if n_note:
        ws.cell(r,7).value = n_note; ws.cell(r,7).font = font(BLACK, italic=True, size=9)
        ws.cell(r,7).alignment = left()

    if alt:
        for ci in [1,2,3]: ws.cell(r,ci).fill = fill(LIGHT_BLUE)
        for ci in [5,6,7]: ws.cell(r,ci).fill = fill(LIGHT_TEAL)

dashboard_rows = [
    ("Company",                  "Palantir Technologies",          None,              FMT_TEXT,         "Company",                  "ServiceNow, Inc.",               None,              FMT_TEXT),
    ("Ticker",                   "PLTR",                           None,              FMT_TEXT,         "Ticker",                   "NOW",                            None,              FMT_TEXT),
    ("Current Price ($)",        f"=PLTR_Model!$B${pltr_price_row}", None,            FMT_PRICE,        "Current Price ($)",        f"=NOW_Model!$B${now_price_row}",  None,              FMT_PRICE),
    ("Market Cap ($B, approx.)", f"=PLTR_Model!$B${pltr_price_row}*PLTR_Model!$B${pltr_shares_row}/1000000", "Price x Diluted Shares", FMT_CURRENCY_DEC,
                                                                                                         "Market Cap ($B, approx.)", f"=NOW_Model!$B${now_price_row}*NOW_Model!$B${now_shares_row}/1000000", "Price x Diluted Shares", FMT_CURRENCY_DEC),
    ("FY2026 Revenue ($M, Base)", f"=PLTR_Model!B{pltr_rev_rows['Base']}", "Base scenario", FMT_CURRENCY, "FY2026 Revenue ($M, Base)", f"=NOW_Model!B{now_rev_rows['Base']}", "Base scenario", FMT_CURRENCY),
    ("FY2026 Rev Growth (Base)", f"=PLTR_Model!B{pltr_growth_rows['Base']}", "Base scenario", FMT_PCT,   "FY2026 Rev Growth (Base)", f"=NOW_Model!B{now_growth_rows['Base']}", "Base scenario", FMT_PCT),
    ("Gross Margin (Q1 2026)",   0.87,                             "Q1 2026 actual",  FMT_PCT,          "Gross Margin (Q1 2026)",   0.795,                            "Non-GAAP Q1 2026", FMT_PCT),
    ("Adj. Op Margin (Q1 2026)", 0.60,                             "Adj./Non-GAAP",   FMT_PCT,          "Non-GAAP Op Margin (Q1 '26)", 0.32,                          "Non-GAAP Q1 2026", FMT_PCT),
    ("FCF Margin (Q1 2026)",     0.57,                             "Adj. FCF margin", FMT_PCT,          "FCF Margin (Q1 2026)",     0.44,                             "Non-GAAP",        FMT_PCT),
    ("Rule of 40 (Q1 2026)",     145,                              "85%+60%=145",     FMT_RO40,         "Rule of 40 (Q1 2026)",     54,                               "22%+32%=54",      FMT_RO40),
    ("EV/Sales FY2026 (est.)",   46.5,                             "Analyst est.",    FMT_MULT,         "EV/Sales FY2026 (est.)",   7.5,                              "Analyst est.",    FMT_MULT),
    ("EV/FCF FY2026 (est.)",     82.7,                             "Analyst est.",    FMT_MULT,         "EV/FCF FY2026 (est.)",     21.3,                             "Analyst est.",    FMT_MULT),
    ("DCF Intrinsic — Bull ($)", f"=PLTR_Model!$B${pltr_dcf_rows['Bull']}", "Linked from model", FMT_PRICE, "DCF Intrinsic — Bull ($)", f"=NOW_Model!$B${now_dcf_rows['Bull']}", "Linked from model", FMT_PRICE),
    ("DCF Intrinsic — Base ($)", f"=PLTR_Model!$B${pltr_dcf_rows['Base']}", "Linked from model", FMT_PRICE, "DCF Intrinsic — Base ($)", f"=NOW_Model!$B${now_dcf_rows['Base']}", "Linked from model", FMT_PRICE),
    ("DCF Intrinsic — Bear ($)", f"=PLTR_Model!$B${pltr_dcf_rows['Bear']}", "Linked from model", FMT_PRICE, "DCF Intrinsic — Bear ($)", f"=NOW_Model!$B${now_dcf_rows['Bear']}", "Linked from model", FMT_PRICE),
    ("Price vs DCF Base (%)",    f"=(PLTR_Model!$B${pltr_price_row}-PLTR_Model!$B${pltr_dcf_rows['Base']})/PLTR_Model!$B${pltr_dcf_rows['Base']}", "(-) = premium", FMT_PCT,
                                                                                                         "Price vs DCF Base (%)",    f"=(NOW_Model!$B${now_price_row}-NOW_Model!$B${now_dcf_rows['Base']})/NOW_Model!$B${now_dcf_rows['Base']}", "(-) = premium", FMT_PCT),
    ("IV (%)",                   0.6879,                           "June 2026",       FMT_PCT,          "IV (%, estimate)",         None,                             "Verify live",     FMT_PCT),
    ("IV Rank",                  46.73,                            "June 2026",       "#,##0.0",        "IV Rank (estimate)",       None,                             "Verify live",     "#,##0.0"),
    ("200-day SMA ($)",          162.0,                            "June 2026",       FMT_PRICE,        "52-wk High ($)",           211.48,                           "52-wk range",     FMT_PRICE),
    ("52-Week High (est., $)",   None,                             "~$170 est.",      FMT_PRICE,        "52-wk Low ($)",            81.24,                            "52-wk range",     FMT_PRICE),
    ("52-Week Low (est., $)",    None,                             "~$70 est.",       FMT_PRICE,        "200-day SMA ($)",          None,                             "Verify live",     FMT_PRICE),
]

for i, row_data in enumerate(dashboard_rows):
    p_label, p_val, p_note, p_fmt, n_label, n_val, n_note, n_fmt = row_data
    dash_row(ws_dash, r, p_label, p_val, p_note, p_fmt, n_label, n_val, n_note, n_fmt, alt=(i%2==0))
    if "DCF Intrinsic" in p_label:
        ws_dash.cell(r,2).fill = fill(LIGHT_GOLD)
        ws_dash.cell(r,6).fill = fill(LIGHT_GOLD)
    r += 1

# Legend
r += 2
style_section_header(ws_dash, r, 1, 7, "COLOR CODING LEGEND", DARK_NAVY)
r += 1
for label, desc, clr in [
    ("Blue text",   "Hardcoded inputs — user-editable assumptions and market data", BLUE),
    ("Black text",  "Formulas and calculated outputs", BLACK),
    ("Green text",  "Cross-sheet links pulling data from PLTR_Model or NOW_Model", GREEN),
    ("Gold background", "Key output cells: DCF intrinsic values, Rule of 40 scores", BLACK),
]:
    ws_dash.cell(r,1).value = label; ws_dash.cell(r,1).font = font(clr, bold=True)
    if "Gold" in label: ws_dash.cell(r,1).fill = fill(GOLD)
    ws_dash.cell(r,2).value = desc; ws_dash.cell(r,2).font = font(BLACK)
    ws_dash.merge_cells(f"B{r}:G{r}")
    r += 1

# ════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════
out_path = "/sessions/ecstatic-upbeat-wright/mnt/Stock Ticker Analysis/PLTR_NOW_Investment_Model.xlsx"
wb.save(out_path)
print("SAVED:", out_path)
print("PLTR DCF rows:", pltr_dcf_rows)
print("NOW  DCF rows:", now_dcf_rows)
print("PLTR price row:", pltr_price_row)
print("NOW  price row:", now_price_row)
