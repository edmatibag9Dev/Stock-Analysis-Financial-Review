"""
Sweetgreen (SG) Investment Model Builder
Builds a professional 4-sheet Excel model using openpyxl
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# Color constants
NAVY        = "1F4E79"
BLUE_INPUT  = "0000FF"
BLACK       = "000000"
GREEN_LINK  = "008000"
WHITE       = "FFFFFF"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GRAY  = "F2F2F2"
DARK_GRAY   = "595959"
AMBER       = "FFC000"
RED         = "FF0000"
GREEN_FILL  = "E2EFDA"
NAVY_LIGHT  = "BDD7EE"

FMT_CURR_DEC  = '$#,##0.0'
FMT_PCT       = '0.0%'
FMT_MULT      = '0.00"x"'
FMT_ZERO_DASH = '#,##0;(#,##0);"-"'
FMT_CURR_DASH = '$#,##0;($#,##0);"-"'
FMT_PCT_DASH  = '0.0%;(0.0%);"-"'
FMT_NUM2      = '#,##0.0;(#,##0.0);"-"'


def font(bold=False, size=11, color=BLACK, name="Arial", italic=False):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def set_cell(ws, row, col, value, bold=False, color=BLACK, fill_color=None,
             h_align="left", num_format=None, border=None, size=11,
             italic=False, wrap=False, comment_text=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    c.font = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(h=h_align, wrap=wrap)
    if fill_color:
        c.fill = fill(fill_color)
    if num_format:
        c.number_format = num_format
    if border:
        c.border = border
    if comment_text:
        cmt = Comment(comment_text, "Model Builder")
        cmt.width = 300
        cmt.height = 100
        c.comment = cmt
    return c


def header_cell(ws, row, col, text, span=None, fill_color=NAVY, text_color=WHITE,
                size=13, h_align="center"):
    c = set_cell(ws, row, col, text, bold=True, color=text_color,
                 fill_color=fill_color, h_align=h_align, size=size)
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return c


def col_header(ws, row, col, text, fill_color=NAVY_LIGHT):
    return set_cell(ws, row, col, text, bold=True, fill_color=fill_color,
                    h_align="center", border=border_thin())


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: Dashboard
# ════════════════════════════════════════════════════════════════════════════

def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 2

    header_cell(ws, 1, 2, "SWEETGREEN (SG) — INVESTMENT DASHBOARD",
                span=3, fill_color=NAVY, text_color=WHITE, size=16, h_align="center")
    ws.row_dimensions[1].height = 32
    set_cell(ws, 2, 2, "As of June 5, 2026  |  Source: Q1 2026 8-K, SEC EDGAR CIK 0001477815, Finviz",
             italic=True, color=DARK_GRAY, size=10)

    def sec_hdr(row, text):
        header_cell(ws, row, 2, text, span=3, fill_color="2E75B6",
                    text_color=WHITE, size=11, h_align="left")
        ws.row_dimensions[row].height = 20

    sec_hdr(4, "  COMPANY SNAPSHOT")

    col_header(ws, 5, 2, "Metric")
    col_header(ws, 5, 3, "Value")
    col_header(ws, 5, 4, "Source / Notes")

    snap_rows = [
        ("Current Price",                "$7.42",          "June 5, 2026",                    None,          BLUE_INPUT, "Market price June 5 2026"),
        ("52-Week Range",                "$4.49 - $16.70", "Finviz June 5 2026",              None,          BLUE_INPUT, "52-week high/low Finviz"),
        ("Market Cap (~$M)",             881,              "118.8M shares x $7.42",           FMT_CURR_DASH, BLUE_INPUT, "118.8M diluted shares x $7.42"),
        ("Enterprise Value (~$M)",       720,              "Mkt Cap minus $161M net cash",    FMT_CURR_DASH, BLUE_INPUT, "~$881M mkt cap minus $161M net cash"),
        ("EV/FY2025 Revenue",            1.06,             "$720M / $679.5M",                 FMT_MULT,      BLUE_INPUT, "EV divided by FY2025 revenue"),
        ("FY2025 Revenue ($M)",          679.5,            "FY2025 8-K",                      FMT_NUM2,      BLUE_INPUT, "Full-year 2025 revenue annual 8-K"),
        ("Q1 2026 Revenue ($M)",         161.5,            "Q1 2026 8-K",                     FMT_NUM2,      BLUE_INPUT, "Q1 2026 revenue May 7 2026 8-K"),
        ("Q1 2026 SSS Change",           -0.128,           "Q1 2026 8-K",                     FMT_PCT,       BLUE_INPUT, "Same-store sales Q1 2026"),
        ("Restaurant-Level Margin Q1",   0.100,            "Q1 2026 8-K",                     FMT_PCT,       BLUE_INPUT, "Rest-level profit margin Q1 2026"),
        ("Adj. EBITDA Q1 2026 ($M)",     -8.1,             "Q1 2026 8-K",                     FMT_NUM2,      BLUE_INPUT, "Adjusted EBITDA Q1 2026"),
        ("Cash + Restricted ($M)",       160.9,            "Balance sheet",                   FMT_NUM2,      BLUE_INPUT, "Cash plus restricted cash Q1 2026"),
        ("Wonder Group Investment ($M)", 86.4,             "Balance sheet (illiquid)",        FMT_NUM2,      BLUE_INPUT, "Investment in Wonder Group - illiquid"),
        ("Total Restaurants",            285,              "Q1 2026",                         FMT_ZERO_DASH, BLUE_INPUT, "Restaurant count Q1 2026"),
        ("Infinite Kitchen Locations",   33,               "Q1 2026",                         FMT_ZERO_DASH, BLUE_INPUT, "Infinite Kitchen locations Q1 2026"),
        ("AUV (trailing, $K)",           2572,             "Q1 2026",                         FMT_CURR_DASH, BLUE_INPUT, "Annualized unit volume trailing"),
        ("FY2026 SSS Guidance",          "(4%) to (2%)",   "Management guidance",             None,          BLUE_INPUT, "Full-year 2026 SSS guidance"),
        ("FY2026 Adj. EBITDA Guidance",  "$1M to $6M",     "Management guidance",             None,          BLUE_INPUT, "Full-year 2026 Adj EBITDA guidance"),
    ]

    for i, (label, value, source, nfmt, txt_color, cmt) in enumerate(snap_rows):
        r = 6 + i
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        set_cell(ws, r, 2, label, bold=False, fill_color=bg, border=border_thin())
        vc = set_cell(ws, r, 3, value, color=txt_color, fill_color=bg,
                      h_align="center", border=border_thin(), comment_text=cmt)
        if nfmt:
            vc.number_format = nfmt
        set_cell(ws, r, 4, source, italic=True, color=DARK_GRAY,
                 fill_color=bg, border=border_thin(), size=9)

    # DCF cross-sheet link — placeholder, will update after SG_Model is built
    r = 6 + len(snap_rows)
    bg = LIGHT_GRAY if len(snap_rows) % 2 == 0 else WHITE
    set_cell(ws, r, 2, "DCF Bull / Base / Bear ($/share)", bold=True,
             fill_color=bg, border=border_thin())
    lc = ws.cell(row=r, column=3)
    lc.value = "=SG_Model!C999"  # placeholder replaced later
    lc.font = font(bold=True, color=GREEN_LINK)
    lc.alignment = align(h="center")
    lc.fill = fill(bg)
    lc.border = border_thin()
    set_cell(ws, r, 4, "Cross-sheet link to SG_Model DCF prices",
             italic=True, color=GREEN_LINK, fill_color=bg, border=border_thin(), size=9)
    DCF_LINK_ROW = r

    # Technical table
    tr = r + 2
    sec_hdr(tr, "  TECHNICAL PRICE LEVELS (Finviz, June 5, 2026)")
    tr += 1

    col_header(ws, tr, 2, "Indicator")
    col_header(ws, tr, 3, "Price ($)")
    col_header(ws, tr, 4, "Signal")
    tr += 1

    tech_rows = [
        ("SMA 20",  8.35, "Price BELOW 20-day — short-term bearish",    RED),
        ("SMA 50",  7.08, "Price ABOVE 50-day — near-term minor support","2E7D32"),
        ("SMA 200", 8.99, "Price BELOW 200-day — long-term downtrend",  RED),
    ]
    for i, (label, val, sig, sig_color) in enumerate(tech_rows):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        set_cell(ws, tr, 2, label, fill_color=bg, border=border_thin())
        vc = set_cell(ws, tr, 3, val, color=BLUE_INPUT, fill_color=bg,
                      h_align="center", border=border_thin(),
                      comment_text="SMA data from Finviz June 5 2026")
        vc.number_format = FMT_CURR_DEC
        set_cell(ws, tr, 4, sig, color=sig_color, fill_color=bg,
                 border=border_thin(), size=9, italic=True)
        tr += 1

    set_cell(ws, tr, 2, "Overall Technical Status", bold=True,
             fill_color=LIGHT_BLUE, border=border_thin())
    set_cell(ws, tr, 3, "BEARISH", bold=True, color=RED,
             fill_color=LIGHT_BLUE, h_align="center", border=border_thin())
    set_cell(ws, tr, 4, "Below 20-day and 200-day SMAs — cautious near-term",
             italic=True, color=DARK_GRAY, fill_color=LIGHT_BLUE,
             border=border_thin(), size=9)
    tr += 2

    # Investment thesis
    sec_hdr(tr, "  INVESTMENT THESIS SUMMARY")
    tr += 1
    thesis = [
        ("BULL CASE", "IK rollout drives margin recovery. New mgmt restores comps. Target: ~$11-12/share.", NAVY, "1F4E79"),
        ("BASE CASE", "SSS stabilizes at guidance. Slow margin recovery. Target: ~$5.47/share.", DARK_GRAY, "595959"),
        ("BEAR CASE", "Comp erosion accelerates. Cash burn risk. Dilution risk. Target: ~$1.52/share.", RED, "C00000"),
    ]
    for case, text, color, fill_c in thesis:
        set_cell(ws, tr, 2, case, bold=True, color=WHITE, fill_color=fill_c, border=border_thin())
        set_cell(ws, tr, 3, text, fill_color=LIGHT_GRAY, border=border_thin(), size=9, wrap=True)
        ws.merge_cells(start_row=tr, start_column=3, end_row=tr, end_column=4)
        ws.row_dimensions[tr].height = 36
        tr += 1

    for rn in range(6, 6 + len(snap_rows) + 1):
        ws.row_dimensions[rn].height = 18

    return ws, DCF_LINK_ROW


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: SG_Model
# ════════════════════════════════════════════════════════════════════════════

def build_sg_model(wb):
    ws = wb.create_sheet("SG_Model")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14

    COL_LABEL = 2
    row = 1

    # Title
    header_cell(ws, row, 2, "SWEETGREEN (SG) — FINANCIAL MODEL",
                span=8, fill_color=NAVY, text_color=WHITE, size=15, h_align="center")
    ws.row_dimensions[row].height = 30
    row += 1
    set_cell(ws, row, 2,
             "Source: Sweetgreen Q1 2026 8-K (May 7, 2026), Q4 2025 8-K (Feb 26, 2026), SEC EDGAR CIK 0001477815",
             italic=True, color=DARK_GRAY, size=9)
    row += 2

    # ── SECTION A ────────────────────────────────────────────────────────────
    header_cell(ws, row, 2, "SECTION A: HISTORICAL FINANCIALS  ($M unless noted)",
                span=8, fill_color="2E75B6", text_color=WHITE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    hist_cols = ["Metric", "FY2024", "FY2025", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Q1 2026"]
    for j, hdr in enumerate(hist_cols):
        col_header(ws, row, COL_LABEL + j, hdr,
                   fill_color=NAVY_LIGHT if j > 0 else NAVY)
    row += 1

    REV_ROW = row
    rev_data = [676.8, 679.5, 166.3, 185.6, 172.4, 155.2, 161.5]
    set_cell(ws, row, COL_LABEL, "Total Revenue ($M)", bold=True, border=border_thin())
    for j, v in enumerate(rev_data):
        c = set_cell(ws, row, COL_LABEL + 1 + j, v, color=BLUE_INPUT,
                     h_align="center", border=border_thin(),
                     comment_text="Source: Sweetgreen SEC filings CIK 0001477815")
        c.number_format = FMT_NUM2
    row += 1

    YOY_ROW = row
    set_cell(ws, row, COL_LABEL, "  YoY Revenue Growth %", italic=True, border=border_thin())
    # FY2024: no prior
    c = ws.cell(row=row, column=COL_LABEL+1)
    c.value = None; c.border = border_thin(); c.alignment = align(h="center")
    # FY2025 vs FY2024
    c = ws.cell(row=row, column=COL_LABEL+2)
    c.value = f"=IF(C{REV_ROW}=0,0,D{REV_ROW}/C{REV_ROW}-1)"
    c.font = font(color=BLACK); c.number_format = FMT_PCT; c.alignment = align(h="center"); c.border = border_thin()
    # Q1-Q4 quarters: no direct YoY in this layout
    for j in range(3, 7):
        c = ws.cell(row=row, column=COL_LABEL+1+j)
        c.value = None; c.border = border_thin(); c.alignment = align(h="center")
    # Q1 2026 vs Q1 2025
    c = ws.cell(row=row, column=COL_LABEL+7)
    c.value = f"=IF(E{REV_ROW}=0,0,I{REV_ROW}/E{REV_ROW}-1)"
    c.font = font(color=BLACK); c.number_format = FMT_PCT; c.alignment = align(h="center"); c.border = border_thin()
    row += 1

    SSS_DATA = [0.062, -0.079, -0.031, -0.050, -0.080, -0.115, -0.128]
    set_cell(ws, row, COL_LABEL, "Same-Store Sales Change %", border=border_thin())
    for j, v in enumerate(SSS_DATA):
        c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin(),
                     comment_text="Source: Sweetgreen quarterly/annual SEC filings")
        c.number_format = FMT_PCT
    row += 1

    RLP_ROW = row
    RLP_DATA = [114, 82, 29.7, 25, 20, 16.2, 16.2]
    set_cell(ws, row, COL_LABEL, "Restaurant-Level Profit ($M)", border=border_thin())
    for j, v in enumerate(RLP_DATA):
        c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin(),
                     comment_text="Source: Sweetgreen SEC filings - approx for unreported quarters")
        c.number_format = FMT_NUM2
    row += 1

    set_cell(ws, row, COL_LABEL, "  Restaurant-Level Margin %", italic=True, border=border_thin())
    for j in range(7):
        col_l = get_column_letter(COL_LABEL+1+j)
        c = ws.cell(row=row, column=COL_LABEL+1+j)
        c.value = f"=IF({col_l}{REV_ROW}=0,0,{col_l}{RLP_ROW}/{col_l}{REV_ROW})"
        c.font = font(color=BLACK); c.number_format = FMT_PCT; c.alignment = align(h="center"); c.border = border_thin()
    row += 1

    AEBITDA_ROW = row
    AEBITDA_DATA = [0, -20, 0.3, -3, -4, -13.3, -8.1]
    set_cell(ws, row, COL_LABEL, "Adj. EBITDA ($M)", border=border_thin())
    for j, v in enumerate(AEBITDA_DATA):
        c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin(),
                     comment_text="Source: Sweetgreen earnings releases and 8-K filings")
        c.number_format = FMT_NUM2
    row += 1

    set_cell(ws, row, COL_LABEL, "  Adj. EBITDA Margin %", italic=True, border=border_thin())
    for j in range(7):
        col_l = get_column_letter(COL_LABEL+1+j)
        c = ws.cell(row=row, column=COL_LABEL+1+j)
        c.value = f"=IF({col_l}{REV_ROW}=0,0,{col_l}{AEBITDA_ROW}/{col_l}{REV_ROW})"
        c.font = font(color=BLACK); c.number_format = FMT_PCT; c.alignment = align(h="center"); c.border = border_thin()
    row += 1

    REST_DATA = [246, 281, 251, 261, 271, 281, 285]
    set_cell(ws, row, COL_LABEL, "Total Restaurants (end of period)", border=border_thin())
    for j, v in enumerate(REST_DATA):
        c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin(),
                     comment_text="Restaurant count from SEC filings")
        c.number_format = FMT_ZERO_DASH
    row += 1

    IK_DATA = [10, 27, 18, 22, 25, 27, 33]
    set_cell(ws, row, COL_LABEL, "Infinite Kitchen Locations", border=border_thin())
    for j, v in enumerate(IK_DATA):
        c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin(),
                     comment_text="Infinite Kitchen automated restaurant count")
        c.number_format = FMT_ZERO_DASH
    row += 1

    AUV_DATA = [2900, 2800, 2907, 2800, 2700, 2600, 2572]
    set_cell(ws, row, COL_LABEL, "AUV ($K, trailing)", border=border_thin())
    for j, v in enumerate(AUV_DATA):
        c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin(),
                     comment_text="Annualized unit volume from Sweetgreen quarterly reports")
        c.number_format = FMT_CURR_DASH
    row += 2

    # ── SECTION B ────────────────────────────────────────────────────────────
    header_cell(ws, row, 2, "SECTION B: REVENUE & EBITDA PROJECTIONS (FY2026-FY2030)  |  THREE SCENARIOS",
                span=8, fill_color="2E75B6", text_color=WHITE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    note_c = ws.cell(row=row, column=COL_LABEL)
    note_c.value = ("NOTE: FY2026 base revenue = $655M (derived from guidance: SSS -4% to -2% + ~13 new restaurants). "
                    "Blue = hardcoded inputs. Black = formulas. Green = cross-sheet links.")
    note_c.font = font(italic=True, size=9, color=DARK_GRAY)
    note_c.alignment = align(wrap=True)
    ws.merge_cells(start_row=row, start_column=COL_LABEL, end_row=row, end_column=COL_LABEL+7)
    ws.row_dimensions[row].height = 28
    row += 1

    # ── Revenue Growth Assumptions
    header_cell(ws, row, 2, "REVENUE GROWTH ASSUMPTIONS (annual %, from $655M FY2026 guidance base — YoY% vs. FY2025 will differ)",
                span=8, fill_color=LIGHT_BLUE, text_color=NAVY, size=10, h_align="left")
    row += 1
    for j, h in enumerate(["Year", "Bull Case", "Base Case", "Bear Case"]):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
    row += 1

    GROWTH_ROWS = {}
    rev_growth_data = [
        ("FY2026", 0.00, -0.03, -0.05),
        ("FY2027", 0.08,  0.04, -0.02),
        ("FY2028", 0.11,  0.04, -0.02),
        ("FY2029", 0.11,  0.04,  0.01),
        ("FY2030", 0.10,  0.04,  0.02),
    ]
    for yr, bull, base, bear in rev_growth_data:
        GROWTH_ROWS[yr] = row
        set_cell(ws, row, COL_LABEL, yr, bold=True, border=border_thin())
        for j, v in enumerate([bull, base, bear]):
            c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin())
            c.number_format = FMT_PCT
        row += 1
    row += 1

    # ── RLM Assumptions
    header_cell(ws, row, 2, "RESTAURANT-LEVEL MARGIN ASSUMPTIONS (%)",
                span=8, fill_color=LIGHT_BLUE, text_color=NAVY, size=10, h_align="left")
    row += 1
    for j, h in enumerate(["Year", "Bull Case", "Base Case", "Bear Case"]):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
    row += 1

    RLM_ROWS = {}
    rlm_data = [
        ("FY2026", 0.145, 0.143, 0.120),
        ("FY2027", 0.180, 0.150, 0.125),
        ("FY2028", 0.210, 0.160, 0.130),
        ("FY2029", 0.230, 0.170, 0.135),
        ("FY2030", 0.250, 0.175, 0.140),
    ]
    for yr, bull, base, bear in rlm_data:
        RLM_ROWS[yr] = row
        set_cell(ws, row, COL_LABEL, yr, bold=True, border=border_thin())
        for j, v in enumerate([bull, base, bear]):
            c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin())
            c.number_format = FMT_PCT
        row += 1
    row += 1

    # ── G&A Assumptions
    header_cell(ws, row, 2, "G&A AS % OF REVENUE (declining with scale)",
                span=8, fill_color=LIGHT_BLUE, text_color=NAVY, size=10, h_align="left")
    row += 1
    for j, h in enumerate(["Year", "Bull Case", "Base Case", "Bear Case"]):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
    row += 1

    GA_ROWS = {}
    ga_data = [
        ("FY2026", 0.17, 0.18, 0.19),
        ("FY2027", 0.15, 0.17, 0.19),
        ("FY2028", 0.13, 0.16, 0.18),
        ("FY2029", 0.12, 0.15, 0.18),
        ("FY2030", 0.11, 0.14, 0.17),
    ]
    for yr, bull, base, bear in ga_data:
        GA_ROWS[yr] = row
        set_cell(ws, row, COL_LABEL, yr, bold=True, border=border_thin())
        for j, v in enumerate([bull, base, bear]):
            c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, h_align="center", border=border_thin())
            c.number_format = FMT_PCT
        row += 1
    row += 1

    # Base revenue anchor
    BASE_REV_ROW = row
    set_cell(ws, row, COL_LABEL, "FY2026 Base Starting Revenue ($M):", bold=True, border=border_thin())
    c = set_cell(ws, row, COL_LABEL+1, 655, color=BLUE_INPUT, h_align="center", border=border_thin(),
                 comment_text="FY2026 base revenue derived from management SSS guidance (-4% to -2%) plus ~13 net new restaurants partial-year contribution")
    c.number_format = FMT_NUM2
    BASE_REV_CELL = f"{get_column_letter(COL_LABEL+1)}{row}"
    row += 2

    # ── Projection tables for each scenario
    years = ["FY2026","FY2027","FY2028","FY2029","FY2030"]
    scenarios = [("Bull", 1, "1F7A1F"), ("Base", 2, NAVY), ("Bear", 3, "C00000")]

    # Store key cell addresses
    rev_cells    = {}
    rlp_cells    = {}
    ga_cells     = {}
    ebitda_cells = {}
    ebitda_rc    = {}  # {scenario: {year: (row, col)}}

    for sc_name, s_off, s_fill in scenarios:
        rev_cells[sc_name]    = {}
        rlp_cells[sc_name]    = {}
        ga_cells[sc_name]     = {}
        ebitda_cells[sc_name] = {}
        ebitda_rc[sc_name]    = {}

        header_cell(ws, row, 2, f"  PROJECTED FINANCIALS - {sc_name.upper()} CASE ($M)",
                    span=8, fill_color=s_fill, text_color=WHITE, size=10, h_align="left")
        row += 1
        for j, h in enumerate(["Metric"] + years):
            col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
        row += 1

        # Revenue
        REV_PROJ_ROW = row
        set_cell(ws, row, COL_LABEL, "Revenue ($M)", bold=True, border=border_thin())
        g_col = get_column_letter(COL_LABEL + s_off)
        for j, yr in enumerate(years):
            col = COL_LABEL+1+j
            col_l = get_column_letter(col)
            gr = GROWTH_ROWS[yr]
            if j == 0:
                formula = f"={BASE_REV_CELL}*(1+{g_col}{gr})"
            else:
                prev_l = get_column_letter(COL_LABEL+j)
                formula = f"={prev_l}{REV_PROJ_ROW}*(1+{g_col}{gr})"
            c = ws.cell(row=row, column=col)
            c.value = formula
            c.font = font(color=BLACK); c.number_format = FMT_NUM2
            c.alignment = align(h="center"); c.border = border_thin()
            rev_cells[sc_name][yr] = f"{col_l}{row}"
        row += 1

        # YoY Growth
        set_cell(ws, row, COL_LABEL, "  YoY Growth %", italic=True, border=border_thin())
        fy25_col = get_column_letter(COL_LABEL+2)  # FY2025 historical col
        for j, yr in enumerate(years):
            col = COL_LABEL+1+j
            col_l = get_column_letter(col)
            c = ws.cell(row=row, column=col)
            if j == 0:
                c.value = f"=IF({fy25_col}{REV_ROW}=0,0,{col_l}{REV_PROJ_ROW}/{fy25_col}{REV_ROW}-1)"
            else:
                prev_l = get_column_letter(COL_LABEL+j)
                c.value = f"=IF({prev_l}{REV_PROJ_ROW}=0,0,{col_l}{REV_PROJ_ROW}/{prev_l}{REV_PROJ_ROW}-1)"
            c.font = font(color=BLACK); c.number_format = FMT_PCT
            c.alignment = align(h="center"); c.border = border_thin()
        row += 1

        # Restaurant-Level Profit
        RLP_PROJ_ROW = row
        set_cell(ws, row, COL_LABEL, "Restaurant-Level Profit ($M)", border=border_thin())
        m_col = get_column_letter(COL_LABEL + s_off)
        for j, yr in enumerate(years):
            col = COL_LABEL+1+j
            col_l = get_column_letter(col)
            mr = RLM_ROWS[yr]
            rev_a = rev_cells[sc_name][yr]
            c = ws.cell(row=row, column=col)
            c.value = f"={rev_a}*{m_col}{mr}"
            c.font = font(color=BLACK); c.number_format = FMT_NUM2
            c.alignment = align(h="center"); c.border = border_thin()
            rlp_cells[sc_name][yr] = f"{col_l}{row}"
        row += 1

        # RL Margin display
        set_cell(ws, row, COL_LABEL, "  Restaurant-Level Margin %", italic=True, border=border_thin())
        for j, yr in enumerate(years):
            col = COL_LABEL+1+j
            mr = RLM_ROWS[yr]
            c = ws.cell(row=row, column=col)
            c.value = f"={m_col}{mr}"
            c.font = font(color=BLACK); c.number_format = FMT_PCT
            c.alignment = align(h="center"); c.border = border_thin()
        row += 1

        # G&A
        GA_PROJ_ROW = row
        set_cell(ws, row, COL_LABEL, "G&A Expense ($M)", border=border_thin())
        ga_col = get_column_letter(COL_LABEL + s_off)
        for j, yr in enumerate(years):
            col = COL_LABEL+1+j
            col_l = get_column_letter(col)
            gar = GA_ROWS[yr]
            rev_a = rev_cells[sc_name][yr]
            c = ws.cell(row=row, column=col)
            c.value = f"={rev_a}*{ga_col}{gar}"
            c.font = font(color=BLACK); c.number_format = FMT_NUM2
            c.alignment = align(h="center"); c.border = border_thin()
            ga_cells[sc_name][yr] = f"{col_l}{row}"
        row += 1

        # G&A %
        set_cell(ws, row, COL_LABEL, "  G&A % of Revenue", italic=True, border=border_thin())
        for j, yr in enumerate(years):
            col = COL_LABEL+1+j
            gar = GA_ROWS[yr]
            c = ws.cell(row=row, column=col)
            c.value = f"={ga_col}{gar}"
            c.font = font(color=BLACK); c.number_format = FMT_PCT
            c.alignment = align(h="center"); c.border = border_thin()
        row += 1

        # Adj. EBITDA
        AEBITDA_PROJ_ROW = row
        set_cell(ws, row, COL_LABEL, "Adj. EBITDA ($M)", bold=True, border=border_thin())
        for j, yr in enumerate(years):
            col = COL_LABEL+1+j
            col_l = get_column_letter(col)
            rlp_a = rlp_cells[sc_name][yr]
            ga_a  = ga_cells[sc_name][yr]
            c = ws.cell(row=row, column=col)
            c.value = f"={rlp_a}-{ga_a}"
            c.font = font(bold=True, color=BLACK); c.number_format = FMT_NUM2
            c.alignment = align(h="center"); c.border = border_thin()
            c.fill = fill(GREEN_FILL)
            ebitda_cells[sc_name][yr] = f"{col_l}{row}"
            ebitda_rc[sc_name][yr] = (row, col)
        row += 1

        # EBITDA Margin %
        set_cell(ws, row, COL_LABEL, "  Adj. EBITDA Margin %", italic=True, border=border_thin())
        for j, yr in enumerate(years):
            col = COL_LABEL+1+j
            col_l = get_column_letter(col)
            ea = ebitda_cells[sc_name][yr]
            ra = rev_cells[sc_name][yr]
            c = ws.cell(row=row, column=col)
            c.value = f"=IF({ra}=0,0,{ea}/{ra})"
            c.font = font(color=BLACK); c.number_format = FMT_PCT
            c.alignment = align(h="center"); c.border = border_thin()
        row += 2

    # ── SECTION C: DCF ────────────────────────────────────────────────────────
    SECT_C = row
    header_cell(ws, row, 2, "SECTION C: DCF VALUATION (EV/EBITDA Terminal Multiple Approach)",
                span=8, fill_color="2E75B6", text_color=WHITE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    note_c = ws.cell(row=row, column=COL_LABEL)
    note_c.value = ("NOTE: Standard FCF DCF not used - company is pre-EBITDA through FY2026. "
                    "Using EV/EBITDA terminal multiple approach with revenue-based cross-check. "
                    "Terminal value based on FY2030 EBITDA x exit multiple, discounted back 5 years.")
    note_c.font = font(italic=True, size=9, color=DARK_GRAY)
    note_c.alignment = align(wrap=True)
    ws.merge_cells(start_row=row, start_column=COL_LABEL, end_row=row, end_column=COL_LABEL+7)
    ws.row_dimensions[row].height = 32
    row += 1

    for j, h in enumerate(["DCF Assumption", "Bull Case", "Base Case", "Bear Case"]):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
    row += 1

    # Discount rate
    DR_ROW = row
    set_cell(ws, row, COL_LABEL, "Discount Rate (WACC)", border=border_thin())
    dr_vals = [0.11, 0.13, 0.15]
    dr_cells = []
    for j, v in enumerate(dr_vals):
        col = COL_LABEL+1+j
        c = set_cell(ws, row, col, v, color=BLUE_INPUT, h_align="center", border=border_thin())
        c.number_format = FMT_PCT
        dr_cells.append(f"{get_column_letter(col)}{row}")
    row += 1

    # Terminal multiple
    TM_ROW = row
    set_cell(ws, row, COL_LABEL, "Terminal EV/EBITDA Multiple", border=border_thin())
    tm_vals = [20, 18, 12]
    tm_cells = []
    for j, v in enumerate(tm_vals):
        col = COL_LABEL+1+j
        c = set_cell(ws, row, col, v, color=BLUE_INPUT, h_align="center", border=border_thin())
        c.number_format = '#,##0"x"'
        tm_cells.append(f"{get_column_letter(col)}{row}")
    row += 1

    # TGR
    TGR_ROW = row
    set_cell(ws, row, COL_LABEL, "Terminal Growth Rate", border=border_thin())
    tgr_vals = [0.03, 0.025, 0.02]
    tgr_cells = []
    for j, v in enumerate(tgr_vals):
        col = COL_LABEL+1+j
        c = set_cell(ws, row, col, v, color=BLUE_INPUT, h_align="center", border=border_thin())
        c.number_format = FMT_PCT
        tgr_cells.append(f"{get_column_letter(col)}{row}")
    row += 1

    # Net cash
    CASH_ROW = row
    set_cell(ws, row, COL_LABEL, "Net Cash ($M)", border=border_thin())
    cash_cells_dcf = []
    for j in range(3):
        col = COL_LABEL+1+j
        c = set_cell(ws, row, col, 161, color=BLUE_INPUT, h_align="center", border=border_thin(),
                     comment_text="Net cash from Q1 2026 balance sheet: $160.9M cash + restricted cash")
        c.number_format = FMT_NUM2
        cash_cells_dcf.append(f"{get_column_letter(col)}{row}")
    row += 1

    # Diluted shares
    SHR_ROW = row
    set_cell(ws, row, COL_LABEL, "Diluted Shares (M)", border=border_thin())
    shr_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        c = set_cell(ws, row, col, 118.8, color=BLUE_INPUT, h_align="center", border=border_thin(),
                     comment_text="Diluted share count Q1 2026 10-Q")
        c.number_format = FMT_NUM2
        shr_cells.append(f"{get_column_letter(col)}{row}")
    row += 2

    # DCF Output
    header_cell(ws, row, 2, "  DCF VALUATION OUTPUT",
                span=8, fill_color=LIGHT_BLUE, text_color=NAVY, size=10, h_align="left")
    row += 1
    for j, h in enumerate(["Line Item", "Bull Case", "Base Case", "Bear Case"]):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
    row += 1

    sc_list = ["Bull", "Base", "Bear"]

    # FY2030 EBITDA
    FY30_E_ROW = row
    set_cell(ws, row, COL_LABEL, "FY2030 Adj. EBITDA ($M)", border=border_thin())
    fy30_e_cells = []
    for j, sc in enumerate(sc_list):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        er, ec = ebitda_rc[sc]["FY2030"]
        ec_l = get_column_letter(ec)
        c = ws.cell(row=row, column=col)
        c.value = f"={ec_l}{er}"
        c.font = font(color=GREEN_LINK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        fy30_e_cells.append(f"{col_l}{row}")
    row += 1

    # Terminal Value
    TV_ROW = row
    set_cell(ws, row, COL_LABEL, "Terminal Value (EBITDA x Multiple, $M)", border=border_thin())
    tv_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"=MAX(0,{fy30_e_cells[j]})*{tm_cells[j]}"
        c.font = font(color=BLACK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        tv_cells.append(f"{col_l}{row}")
    row += 1

    # PV of Terminal Value
    PVTV_ROW = row
    set_cell(ws, row, COL_LABEL, "PV of Terminal Value ($M, 5-yr discount)", border=border_thin())
    pvtv_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"={tv_cells[j]}/((1+{dr_cells[j]})^5)"
        c.font = font(color=BLACK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        pvtv_cells.append(f"{col_l}{row}")
    row += 1

    # PV interim positive EBITDA
    PVINT_ROW = row
    set_cell(ws, row, COL_LABEL, "PV of Interim Positive EBITDA ($M, yrs 3-5)", border=border_thin())
    pvint_cells = []
    for j, sc in enumerate(sc_list):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        er28, ec28 = ebitda_rc[sc]["FY2028"]
        er29, ec29 = ebitda_rc[sc]["FY2029"]
        er30, ec30 = ebitda_rc[sc]["FY2030"]
        e28 = f"{get_column_letter(ec28)}{er28}"
        e29 = f"{get_column_letter(ec29)}{er29}"
        e30 = f"{get_column_letter(ec30)}{er30}"
        dr = dr_cells[j]
        c = ws.cell(row=row, column=col)
        c.value = (f"=MAX(0,{e28})/((1+{dr})^3)"
                   f"+MAX(0,{e29})/((1+{dr})^4)"
                   f"+MAX(0,{e30})/((1+{dr})^5)")
        c.font = font(color=BLACK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        pvint_cells.append(f"{col_l}{row}")
    row += 1

    # Enterprise Value
    EV_ROW = row
    set_cell(ws, row, COL_LABEL, "Enterprise Value ($M)", bold=True, border=border_thin())
    ev_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"={pvtv_cells[j]}"
        c.font = font(bold=True, color=BLACK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        c.fill = fill(GREEN_FILL)
        ev_cells.append(f"{col_l}{row}")
    row += 1

    # + Net Cash
    EQPRE_ROW = row
    set_cell(ws, row, COL_LABEL, "Plus: Net Cash ($M)", border=border_thin())
    eqpre_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"={ev_cells[j]}+{cash_cells_dcf[j]}"
        c.font = font(color=BLACK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        eqpre_cells.append(f"{col_l}{row}")
    row += 1

    # Equity Value
    EQV_ROW = row
    set_cell(ws, row, COL_LABEL, "Equity Value ($M)", bold=True, border=border_thin())
    eqv_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"={eqpre_cells[j]}"
        c.font = font(bold=True, color=BLACK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        c.fill = fill(GREEN_FILL)
        eqv_cells.append(f"{col_l}{row}")
    row += 1

    # Intrinsic Price Per Share — CRITICAL ROW for Dashboard link
    IPS_ROW = row
    set_cell(ws, row, COL_LABEL, "Intrinsic Price Per Share (DCF)", bold=True,
             color=WHITE, fill_color=NAVY, border=border_thin())
    ips_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"=IF({shr_cells[j]}=0,0,{eqv_cells[j]}/{shr_cells[j]})"
        c.font = font(bold=True, color=WHITE); c.number_format = FMT_CURR_DEC
        c.alignment = align(h="center"); c.fill = fill(NAVY)
        c.border = border_thin()
        ips_cells.append(f"{col_l}{row}")
    row += 1

    # Current price
    CP_ROW = row
    set_cell(ws, row, COL_LABEL, "Current Price", border=border_thin())
    cp_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = set_cell(ws, row, col, 7.42, color=BLUE_INPUT, h_align="center", border=border_thin())
        c.number_format = FMT_CURR_DEC
        cp_cells.append(f"{col_l}{row}")
    row += 1

    # Premium/Discount
    PD_ROW = row
    set_cell(ws, row, COL_LABEL, "Premium/(Discount) to Current", border=border_thin())
    for j in range(3):
        col = COL_LABEL+1+j
        c = ws.cell(row=row, column=col)
        c.value = f"=IF({cp_cells[j]}=0,0,{ips_cells[j]}/{cp_cells[j]}-1)"
        c.font = font(color=BLACK); c.number_format = FMT_PCT
        c.alignment = align(h="center"); c.border = border_thin()
    row += 2

    # ── Revenue Multiple Cross-Check
    header_cell(ws, row, 2, "  REVENUE MULTIPLE CROSS-CHECK",
                span=8, fill_color=LIGHT_BLUE, text_color=NAVY, size=10, h_align="left")
    row += 1
    for j, h in enumerate(["Line Item", "Bull Case", "Base Case", "Bear Case"]):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
    row += 1

    # FY2026E Revenue
    set_cell(ws, row, COL_LABEL, "FY2026E Revenue ($M)", border=border_thin())
    rc_rev_cells = []
    for j, v in enumerate([660, 655, 640]):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = set_cell(ws, row, col, v, color=BLUE_INPUT, h_align="center", border=border_thin())
        c.number_format = FMT_NUM2
        rc_rev_cells.append(f"{col_l}{row}")
    row += 1

    # Terminal Revenue (FY2030 from projections)
    set_cell(ws, row, COL_LABEL, "FY2030 Terminal Revenue ($M)", border=border_thin())
    rc_tr_cells = []
    for j, sc in enumerate(sc_list):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        rev30 = rev_cells[sc]["FY2030"]
        c = ws.cell(row=row, column=col)
        c.value = f"={rev30}"
        c.font = font(color=GREEN_LINK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        rc_tr_cells.append(f"{col_l}{row}")
    row += 1

    # EV/Sales multiple
    set_cell(ws, row, COL_LABEL, "Applied EV/Sales Multiple", border=border_thin())
    rc_mult_cells = []
    for j, v in enumerate([3.0, 1.1, 0.7]):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = set_cell(ws, row, col, v, color=BLUE_INPUT, h_align="center", border=border_thin())
        c.number_format = FMT_MULT
        rc_mult_cells.append(f"{col_l}{row}")
    row += 1

    # Implied EV
    set_cell(ws, row, COL_LABEL, "Implied EV ($M)", border=border_thin())
    rc_ev_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"={rc_tr_cells[j]}*{rc_mult_cells[j]}"
        c.font = font(color=BLACK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        rc_ev_cells.append(f"{col_l}{row}")
    row += 1

    # + Net Cash
    set_cell(ws, row, COL_LABEL, "Plus: Net Cash ($M)", border=border_thin())
    rc_cash_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = set_cell(ws, row, col, 161, color=BLUE_INPUT, h_align="center", border=border_thin())
        c.number_format = FMT_NUM2
        rc_cash_cells.append(f"{col_l}{row}")
    row += 1

    # Equity Value
    set_cell(ws, row, COL_LABEL, "Equity Value ($M)", bold=True, border=border_thin())
    rc_eq_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"={rc_ev_cells[j]}+{rc_cash_cells[j]}"
        c.font = font(bold=True, color=BLACK); c.number_format = FMT_NUM2
        c.alignment = align(h="center"); c.border = border_thin()
        c.fill = fill(GREEN_FILL)
        rc_eq_cells.append(f"{col_l}{row}")
    row += 1

    # Implied Price/Share
    set_cell(ws, row, COL_LABEL, "Implied Price/Share (Rev Multiple Check)", bold=True, border=border_thin())
    rc_ips_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"=IF({shr_cells[j]}=0,0,{rc_eq_cells[j]}/{shr_cells[j]})"
        c.font = font(bold=True, color=BLACK); c.number_format = FMT_CURR_DEC
        c.alignment = align(h="center"); c.border = border_thin()
        c.fill = fill(GREEN_FILL)
        rc_ips_cells.append(f"{col_l}{row}")
    row += 1

    # Current Price
    set_cell(ws, row, COL_LABEL, "Current Price", border=border_thin())
    rc_cp_cells = []
    for j in range(3):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = set_cell(ws, row, col, 7.42, color=BLUE_INPUT, h_align="center", border=border_thin())
        c.number_format = FMT_CURR_DEC
        rc_cp_cells.append(f"{col_l}{row}")
    row += 1

    # Premium/Discount
    set_cell(ws, row, COL_LABEL, "Premium/(Discount) to Current", border=border_thin())
    for j in range(3):
        col = COL_LABEL+1+j
        c = ws.cell(row=row, column=col)
        c.value = f"=IF({rc_cp_cells[j]}=0,0,{rc_ips_cells[j]}/{rc_cp_cells[j]}-1)"
        c.font = font(color=BLACK); c.number_format = FMT_PCT
        c.alignment = align(h="center"); c.border = border_thin()
    row += 1

    return ws, IPS_ROW


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: Unit_Economics
# ════════════════════════════════════════════════════════════════════════════

def build_unit_economics(wb):
    ws = wb.create_sheet("Unit_Economics")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 2
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 22
    ws.column_dimensions["L"].width = 28

    COL_LABEL = 2
    row = 1

    header_cell(ws, row, 2, "UNIT ECONOMICS SCORECARD - SG vs. PEERS",
                span=10, fill_color=NAVY, text_color=WHITE, size=15, h_align="center")
    ws.row_dimensions[row].height = 30
    row += 1
    set_cell(ws, row, 2, "Source: Company filings, equity research estimates. Peer data approximate as of June 2026.",
             italic=True, color=DARK_GRAY, size=9)
    row += 2

    header_cell(ws, row, 2, "  SG UNIT ECONOMICS TREND",
                span=5, fill_color="2E75B6", text_color=WHITE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    for j, h in enumerate(["Metric", "FY2024", "FY2025", "Q1 2026", "FY2026E (guidance)"]):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT if j > 0 else NAVY)
    row += 1

    REST_COUNT_ROW = None
    IK_COUNT_ROW   = None

    trend_data = [
        ("AUV ($K, trailing)",           [2900, 2800, 2572, 2600],    FMT_CURR_DASH),
        ("Restaurant-Level Margin",      [0.170, 0.120, 0.100, 0.145], FMT_PCT),
        ("Same-Store Sales %",           [0.062,-0.079,-0.128,-0.030], FMT_PCT),
        ("Total Restaurants",            [246,   281,   285,   294],   FMT_ZERO_DASH),
        ("Infinite Kitchen Locations",   [10,    27,    33,    40],    FMT_ZERO_DASH),
    ]

    for i, (label, vals, nfmt) in enumerate(trend_data):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        set_cell(ws, row, COL_LABEL, label, fill_color=bg, border=border_thin())
        for j, v in enumerate(vals):
            c = set_cell(ws, row, COL_LABEL+1+j, v, color=BLUE_INPUT, fill_color=bg,
                         h_align="center", border=border_thin(),
                         comment_text="Source: Sweetgreen SEC filings and earnings releases")
            c.number_format = nfmt
        if "Total Rest" in label:
            REST_COUNT_ROW = row
        if "Infinite Kitchen Loc" in label:
            IK_COUNT_ROW = row
        row += 1

    # IK % formula row
    bg = WHITE
    set_cell(ws, row, COL_LABEL, "  IK % of Total Restaurants", italic=True, fill_color=bg, border=border_thin())
    for j in range(4):
        col = COL_LABEL+1+j
        col_l = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"=IF({col_l}{REST_COUNT_ROW}=0,0,{col_l}{IK_COUNT_ROW}/{col_l}{REST_COUNT_ROW})"
        c.font = font(color=BLACK); c.number_format = FMT_PCT
        c.alignment = align(h="center"); c.fill = fill(bg); c.border = border_thin()
    row += 1

    # Adj EBITDA
    bg = LIGHT_GRAY
    set_cell(ws, row, COL_LABEL, "Adj. EBITDA ($M)", fill_color=bg, border=border_thin())
    ae_vals = [0, -20, -8.1, None]
    ae_text = ["~$0", "~($20)", "($8.1M)", "$1-$6 guidance"]
    for j in range(4):
        col = COL_LABEL+1+j
        if ae_vals[j] is not None:
            c = set_cell(ws, row, col, ae_vals[j], color=BLUE_INPUT, fill_color=bg,
                         h_align="center", border=border_thin())
            c.number_format = FMT_NUM2
        else:
            set_cell(ws, row, col, ae_text[j], color=BLUE_INPUT, fill_color=bg,
                     h_align="center", border=border_thin(), size=9)
    row += 1

    bg = WHITE
    set_cell(ws, row, COL_LABEL, "Cash Burn (quarterly FCF)", fill_color=bg, border=border_thin())
    for j in range(4):
        val = "($29.6M)" if j == 2 else ""
        c_color = RED if j == 2 else DARK_GRAY
        set_cell(ws, row, COL_LABEL+1+j, val, color=c_color, fill_color=bg,
                 h_align="center", border=border_thin(), size=9,
                 comment_text="Q1 2026 free cash flow from ops minus capex" if j == 2 else None)
    row += 1

    bg = LIGHT_GRAY
    set_cell(ws, row, COL_LABEL, "Implied Cash Runway", fill_color=bg, border=border_thin())
    for j in range(4):
        val = "~5-6 quarters" if j == 3 else ""
        set_cell(ws, row, COL_LABEL+1+j, val, fill_color=bg, h_align="center",
                 border=border_thin(), size=9, italic=(val != ""))
    row += 2

    # ── Peer Comps
    header_cell(ws, row, 2, "  PEER RESTAURANT COMPS (Estimates, June 2026 — verify before trading)",
                span=10, fill_color="2E75B6", text_color=WHITE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    peer_hdrs = ["Company", "Ticker", "Fwd Rev Growth", "Rest.-Level Margin",
                 "EV/Sales", "Market Position", "Note"]
    for j, h in enumerate(peer_hdrs):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
    row += 1

    peers = [
        ("Sweetgreen",  "SG",   "(2%) to 0%", "14-15%", "1.1x",  "Fast-casual salad",        "Declining comps; only peer with negative SSS and sub-15% margins", LIGHT_BLUE, True),
        ("CAVA Group",  "CAVA", "~30%",        "~23%",   "~7x",   "Fast-casual Mediterranean","Best comps in sector; premium valuation justified",                  WHITE,      False),
        ("Chipotle",    "CMG",  "~10%",        "~27%",   "~5.5x", "Fast-casual Mexican",      "Category leader; best unit economics in QSR",                        LIGHT_GRAY, False),
        ("Shake Shack", "SHAK", "~10%",        "~20%",   "~1.5x", "Premium burgers",          "Similar valuation stage; better margins than SG",                   WHITE,      False),
        ("Dutch Bros",  "BROS", "~20%",        "N/A",    "~2.5x", "Drive-thru coffee",        "High growth; asset-light model",                                     LIGHT_GRAY, False),
        ("Wingstop",    "WING", "~15%",        "N/A",    "~4x",   "Wings delivery",           "Asset-light; highest EBITDA margins in group",                       WHITE,      False),
        ("First Watch", "FWRK", "~12%",        "~18%",   "~1.2x", "Daytime dining",           "Similar size and valuation; positive comparable sales",              LIGHT_GRAY, False),
    ]
    for name, ticker, growth, margin, ev_s, pos, note, bg, is_sg in peers:
        data = [name, ticker, growth, margin, ev_s, pos, note]
        for j, v in enumerate(data):
            set_cell(ws, row, COL_LABEL+j, v, bold=is_sg, fill_color=bg,
                     color=BLUE_INPUT if is_sg else BLACK,
                     border=border_thin(), size=9, wrap=(j == 6))
        ws.row_dimensions[row].height = 28 if len(note) > 40 else 18
        row += 1

    row += 1
    note_c = ws.cell(row=row, column=COL_LABEL)
    note_c.value = ("ANALYST NOTE: SG trades at 1.1x EV/Sales — the cheapest of peers on revenue multiple. "
                    "However, it is the ONLY peer with negative same-store sales and sub-15% restaurant-level margins. "
                    "On a risk-adjusted basis, the valuation is not compelling vs. peers with better unit economics. "
                    "Catalyst required: Infinite Kitchen margin recovery or SSS stabilization.")
    note_c.font = font(italic=True, size=9, color=DARK_GRAY)
    note_c.alignment = align(wrap=True)
    ws.merge_cells(start_row=row, start_column=COL_LABEL, end_row=row, end_column=COL_LABEL+7)
    ws.row_dimensions[row].height = 48

    return ws


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: Options_Strategy
# ════════════════════════════════════════════════════════════════════════════

def build_options_strategy(wb):
    ws = wb.create_sheet("Options_Strategy")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 40

    COL_LABEL = 2
    row = 1

    header_cell(ws, row, 2, "OPTIONS STRATEGY - SWEETGREEN (SG)",
                span=6, fill_color=NAVY, text_color=WHITE, size=15, h_align="center")
    ws.row_dimensions[row].height = 30
    row += 1
    set_cell(ws, row, 2,
             "WARNING: Options carry significant risk. Verify current options chain before trading. IV estimates are approximations only.",
             italic=True, color=RED, size=9, bold=True)
    row += 2

    header_cell(ws, row, 2, "  MARKET DATA REFERENCE",
                span=6, fill_color="2E75B6", text_color=WHITE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    mkt_rows = [
        ("Current Price",   "$7.42",       "June 5, 2026"),
        ("52-Week High",    "$16.70",      ""),
        ("52-Week Low",     "$4.49",       ""),
        ("IV (estimated)",  "~85-95%",     "Small-cap speculative estimate - verify current chain"),
        ("IV Rank",         "High (est.)", "Post-run and reversal - premiums elevated"),
        ("User Position",   "None",        "Evaluating entry"),
    ]
    for i, (label, val, note) in enumerate(mkt_rows):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        set_cell(ws, row, COL_LABEL, label, fill_color=bg, border=border_thin())
        set_cell(ws, row, COL_LABEL+1, val, color=BLUE_INPUT, fill_color=bg,
                 h_align="center", border=border_thin(), bold=True)
        set_cell(ws, row, COL_LABEL+2, note, fill_color=bg, border=border_thin(),
                 italic=True, color=DARK_GRAY, size=9)
        row += 1

    row += 1
    header_cell(ws, row, 2, "  RECOMMENDED OPTIONS STRATEGIES",
                span=6, fill_color="2E75B6", text_color=WHITE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    strat_hdrs = ["Strategy", "Structure", "Strikes / Expiry",
                  "Est. Premium/Debit", "Max Gain / Max Loss", "Rationale"]
    for j, h in enumerate(strat_hdrs):
        col_header(ws, row, COL_LABEL+j, h, fill_color=NAVY_LIGHT)
    row += 1

    strategies = [
        {
            "label": "PRIMARY: Cash-Secured Put",
            "bg": GREEN_FILL, "color": "1F7A1F",
            "structure": "Sell 1 put. Set aside Strike x 100 per contract as cash collateral. Net premium collected upfront at trade entry.",
            "strikes": "Sell $6.00 put\nAug 2026 (~10 weeks)",
            "premium": "~$0.80-1.10/contract\n(~13-18% per-cycle yield on capital at risk — verify annualized rate against live options chain)",
            "maxgain": "Max Gain: $80-110/contract (premium kept)\nMax Loss: Assignment at $6.00 minus premium = ~$4.90-5.20 effective basis",
            "rationale": ("High IV inflates put premiums. $4.90-5.20 effective basis is near 52-week low ($4.49). "
                          "Generates income while waiting for better entry. If stock stays above $6, keep premium and repeat. "
                          "Note: effective basis of ~$5.00 is ABOVE base-case DCF (~$3.49) — IK thesis must partially execute to profit."),
        },
        {
            "label": "SPECULATIVE: Bull Call Spread",
            "bg": LIGHT_BLUE, "color": NAVY,
            "structure": "Buy $8 call / Sell $12 call, same expiry. Net debit - defined risk, defined reward. Spreads reduce IV drag vs. outright long calls.",
            "strikes": "Buy $8 Call / Sell $12 Call\nJan 2027 (~7 months)",
            "premium": "~$0.90-1.20 net debit/spread\n($90-120 per spread)",
            "maxgain": "Max Gain: ~$2.80-3.10/spread if stock reaches $12 (~230-260% ROI on spread)\nMax Loss: Debit paid - DEFINED RISK",
            "rationale": ("Defined-risk participation in Infinite Kitchen bull case. $12 target is below prior $16.70 high. "
                          "Only enter with strong conviction on IK thesis. Do not enter without clear catalyst."),
        },
        {
            "label": "AVOID: Long Stock at Market",
            "bg": "FFE0E0", "color": RED,
            "structure": "Buy 100 shares at $7.42 market price. Full equity exposure - no downside protection or premium income.",
            "strikes": "100 shares @ $7.42\nNo expiry",
            "premium": "N/A - full $742 per 100 shares at risk",
            "maxgain": "Max Gain: Unlimited (theoretical)\nMax Loss: -82% in bear case (DCF $1.36) = -$609/100 shares",
            "rationale": ("Base-case DCF is ~$3.49 - less than half current $7.42. Buying at market means pricing in bull-case execution from day one. "
                          "Risk/reward at current price: -53% to base case; -82% to bear case; +99% to bull case. "
                          "Only justified with high-conviction bull-case thesis."),
        },
        {
            "label": "INCOME (IF ASSIGNED): Covered Call",
            "bg": LIGHT_GRAY, "color": DARK_GRAY,
            "structure": "If assigned shares via cash-secured put (~$5.00-5.20 basis), immediately sell covered calls. Roll each expiry to reduce cost basis.",
            "strikes": "Sell $7.00 or $8.00 calls\n6-8 weeks out, rolling",
            "premium": "~$0.50-0.80/contract\nPer 6-8 week rolling cycle",
            "maxgain": "Max Gain: Premium income each cycle + potential upside to strike price\nMax Loss: Existing basis minus all premiums collected over time",
            "rationale": ("If assigned at ~$5.00-5.20, immediately sell covered calls to reduce basis further. "
                          "At ~85% IV, covered call income is substantial. Rolling $7 calls generates ~$0.50-0.80 every 6-8 weeks. "
                          "Continue until stock moves or thesis changes."),
        },
    ]

    for strat in strategies:
        bg = strat["bg"]
        set_cell(ws, row, COL_LABEL,   strat["label"],     bold=True, color=strat["color"],
                 fill_color=bg, border=border_thin(), wrap=True)
        set_cell(ws, row, COL_LABEL+1, strat["structure"],  fill_color=bg, border=border_thin(), wrap=True, size=9)
        set_cell(ws, row, COL_LABEL+2, strat["strikes"],    fill_color=bg, border=border_thin(), wrap=True,
                 size=9, color=BLUE_INPUT)
        set_cell(ws, row, COL_LABEL+3, strat["premium"],    fill_color=bg, border=border_thin(), wrap=True, size=9)
        set_cell(ws, row, COL_LABEL+4, strat["maxgain"],    fill_color=bg, border=border_thin(), wrap=True, size=9)
        set_cell(ws, row, COL_LABEL+5, strat["rationale"],  fill_color=bg, border=border_thin(), wrap=True,
                 size=9, italic=True)
        ws.row_dimensions[row].height = 85
        row += 1

    row += 1
    header_cell(ws, row, 2, "  RISK WARNING",
                span=6, fill_color="C00000", text_color=WHITE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    row += 1

    warn_c = ws.cell(row=row, column=COL_LABEL)
    warn_c.value = ("RISK WARNING: SG is a speculative, pre-profitability stock. "
                    "Position sizing must reflect bear case scenario (-60% from current price). "
                    "Never risk more than you can afford to lose entirely. "
                    "IV estimates of 85-95% are approximations - verify the current options chain before any trade. "
                    "Options can expire worthless. This model is for informational purposes only and does not constitute financial advice.")
    warn_c.font = font(bold=True, size=10, color="C00000")
    warn_c.alignment = align(wrap=True)
    warn_c.fill = fill("FFE0E0")
    warn_c.border = border_thin()
    ws.merge_cells(start_row=row, start_column=COL_LABEL, end_row=row, end_column=COL_LABEL+5)
    ws.row_dimensions[row].height = 72

    return ws


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def build_model():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    print("Building Dashboard...")
    dash_ws, dcf_link_row = build_dashboard(wb)

    print("Building SG_Model...")
    model_ws, ips_row = build_sg_model(wb)

    print("Building Unit_Economics...")
    ue_ws = build_unit_economics(wb)

    print("Building Options_Strategy...")
    opt_ws = build_options_strategy(wb)

    # Fix Dashboard DCF cross-sheet link to actual IPS_ROW
    # IPS_ROW has Bull=C, Base=D, Bear=E  (COL_LABEL+1=C, +2=D, +3=E)
    lc = dash_ws.cell(row=dcf_link_row, column=3)
    bull_col = get_column_letter(3)   # COL_LABEL+1 = col C
    base_col = get_column_letter(4)   # col D
    bear_col = get_column_letter(5)   # col E
    lc.value = (f'=TEXT(SG_Model!{bull_col}{ips_row},"$0.00")'
                f'&" / "&TEXT(SG_Model!{base_col}{ips_row},"$0.00")'
                f'&" / "&TEXT(SG_Model!{bear_col}{ips_row},"$0.00")')
    lc.font = Font(name="Arial", bold=True, color=GREEN_LINK, size=11)
    lc.alignment = Alignment(horizontal="center", vertical="center")

    # Tab colors
    dash_ws.sheet_properties.tabColor  = "1F4E79"
    model_ws.sheet_properties.tabColor = "2E75B6"
    ue_ws.sheet_properties.tabColor    = "70AD47"
    opt_ws.sheet_properties.tabColor   = "FFC000"

    out_path = "/sessions/hopeful-loving-babbage/mnt/Stock Ticker Analysis/SG_Investment_Model.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"IPS_ROW = {ips_row}  (Dashboard links to SG_Model rows C{ips_row}:E{ips_row})")
    return out_path


if __name__ == "__main__":
    build_model()
