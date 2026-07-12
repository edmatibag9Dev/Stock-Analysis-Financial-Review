#!/usr/bin/env python3
"""Build TRMB_Investment_Model.xlsx — Trimble Inc. valuation model.
Brand: teal #2C7A6B, navy #2B4C7E. Data green/red kept separate from brand teal.
Model sheet built first (rows captured), then Dashboard references real cells.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

NAVY="2B4C7E"; TEAL="2C7A6B"; YEL="FFEB9C"; RED="FFC7CE"; GRN="C6EFCE"
BLUE_TXT="0000FF"; GRN_TXT="008000"; RED_TXT="C0392B"; BLACK="000000"
FONT="Inter"
def F(sz=10,b=False,color=BLACK,italic=False): return Font(name=FONT,size=sz,bold=b,color=color,italic=italic)
def fill(c): return PatternFill("solid",fgColor=c)
CUR='$#,##0;($#,##0);"-"'; CUR1='$#,##0.00;($#,##0.00);"-"'
PCT='0.0%;(0.0%);"-"'; MULT='0.0"x"'; CEN=Alignment(horizontal="center")

wb=Workbook()
def hdr(ws,title,ncol):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    c=ws.cell(1,1,title); c.font=F(14,True,"FFFFFF"); c.fill=fill(NAVY)
    c.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=26
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
    s=ws.cell(2,1,"Trimble Inc. (NASDAQ: TRMB)  |  Analyst: Ed  |  Date: 2026-06-11  |  $ in millions unless noted")
    s.font=F(9,False,"FFFFFF",True); s.fill=fill(TEAL); ws.row_dimensions[2].height=16
def sub(ws,row,text,ncol,c=TEAL):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncol)
    cell=ws.cell(row,1,text); cell.font=F(11,True,"FFFFFF"); cell.fill=fill(c)
    cell.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[row].height=20

# ===================================================== MODEL (built first)
ws=wb.create_sheet("Model")
hdr(ws,"TRMB — DCF VALUATION MODEL (Single WACC across all scenarios)",8)
for col,w in zip("ABCDEFGH",[30,13,13,13,13,13,13,13]): ws.column_dimensions[col].width=w
R={}
r=4
inputs=[
 ("price","Current Price ($)",50.00,CUR1,"Market data 2026-06-11 (~$49.90-50.60)"),
 ("shares","Diluted Shares (M)",235.0,'#,##0',"Q1'26 ~236.9M; ~235M guided"),
 ("cash","Cash & Equivalents ($M)",234.1,CUR,"TRMB Q1 2026 8-K Ex-99.1"),
 ("debt","Total Debt ($M)",1412.8,CUR,"TRMB Q1 2026 8-K Ex-99.1"),
 ("netdebt","Net Debt ($M)","=B{debt}-B{cash}",CUR,None),
 ("wacc","WACC (single, all scenarios)",0.095,PCT,"Hybrid SW/HW, net lev 1.1x"),
 ("base25","FY2025E Revenue base ($M)",3520.0,CUR,"Est FY2025 ~$3.48-3.56B guidance"),
 ("rev26g","FY2026E Revenue (guid mid, $M)",3875.0,CUR,"Guidance $3,835-3,915M"),
]
for key,label,v,fmt,note in inputs:
    R[key]=r; ws.cell(r,1,label).font=F(10)
    val=v.format(**R) if isinstance(v,str) else v
    c=ws.cell(r,2,val); c.number_format=fmt
    c.font=F(10,False,BLACK if isinstance(v,str) else BLUE_TXT)
    if note: c.comment=Comment(note,"Model")
    r+=1

r+=1; sub(ws,r,"Scenario Assumptions  (Bull / Base / Bear share ONE WACC = B{})".format(R['wacc']),8); r+=1
ws.cell(r,1,"Assumption").font=F(10,True,"FFFFFF"); ws.cell(r,1).fill=fill(TEAL)
for i,s in enumerate(["Bear","Base","Bull"]):
    c=ws.cell(r,2+i,s); c.font=F(10,True,"FFFFFF"); c.fill=fill([RED_TXT,TEAL,"1E8449"][i]); c.alignment=CEN
r+=1
assump=[("g1","Rev Growth Yr1 (FY26)",[0.07,0.10,0.11]),("g2","Rev Growth Yr2 (FY27)",[0.05,0.09,0.11]),
 ("g3","Rev Growth Yr3 (FY28)",[0.04,0.08,0.10]),("g4","Rev Growth Yr4 (FY29)",[0.04,0.08,0.10]),
 ("g5","Rev Growth Yr5 (FY30)",[0.03,0.07,0.09]),("m1","FCF Margin Yr1",[0.15,0.17,0.18]),
 ("m2","FCF Margin Yr2",[0.155,0.185,0.20]),("m3","FCF Margin Yr3",[0.16,0.20,0.22]),
 ("m4","FCF Margin Yr4",[0.165,0.21,0.235]),("m5","FCF Margin Yr5 (terminal)",[0.17,0.22,0.25]),
 ("tmult","Terminal EV/FCF Multiple",[15.0,20.0,22.0]),("tg","Terminal Growth (x-check)",[0.025,0.035,0.04])]
for key,label,vals in assump:
    R[key]=r; ws.cell(r,1,label).font=F(10)
    for i,v in enumerate(vals):
        c=ws.cell(r,2+i,v); c.font=F(10,False,BLUE_TXT)
        c.number_format=MULT if "Multiple" in label else PCT; c.alignment=CEN
    r+=1

r+=1; sub(ws,r,"Revenue Projection ($M)",8); r+=1
ws.cell(r,1,"Scenario").font=F(10,True)
for i,y in enumerate(["FY26","FY27","FY28","FY29","FY30"]):
    c=ws.cell(r,2+i,y); c.font=F(10,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=CEN
r+=1
revrow={}
for si,sname in enumerate(["Bear","Base","Bull"]):
    col="BCD"[si]; ws.cell(r,1,sname).font=F(10)
    ws.cell(r,2,"=$B${b}*(1+{c}{g1})".format(b=R['base25'],c=col,g1=R['g1'])).number_format=CUR
    ws.cell(r,3,"=B{r}*(1+{c}{g})".format(r=r,c=col,g=R['g2'])).number_format=CUR
    ws.cell(r,4,"=C{r}*(1+{c}{g})".format(r=r,c=col,g=R['g3'])).number_format=CUR
    ws.cell(r,5,"=D{r}*(1+{c}{g})".format(r=r,c=col,g=R['g4'])).number_format=CUR
    ws.cell(r,6,"=E{r}*(1+{c}{g})".format(r=r,c=col,g=R['g5'])).number_format=CUR
    for cc in range(2,7): ws.cell(r,cc).font=F(10)
    revrow[sname]=r; r+=1

r+=1; sub(ws,r,"Free Cash Flow Projection ($M)",8); r+=1
ws.cell(r,1,"Scenario").font=F(10,True)
for i,y in enumerate(["FY26","FY27","FY28","FY29","FY30"]):
    c=ws.cell(r,2+i,y); c.font=F(10,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=CEN
r+=1
fcfrow={}
for si,sname in enumerate(["Bear","Base","Bull"]):
    col="BCD"[si]; rr=revrow[sname]; ws.cell(r,1,sname).font=F(10)
    for ci,mk in enumerate(['m1','m2','m3','m4','m5']):
        cc=ws.cell(r,2+ci,"={cl}{rr}*{c}{m}".format(cl=chr(66+ci),rr=rr,c=col,m=R[mk]))
        cc.number_format=CUR; cc.font=F(10)
    fcfrow[sname]=r; r+=1

r+=1; sub(ws,r,"DCF Valuation Output",8); r+=1
ws.cell(r,1,"Metric").font=F(10,True,"FFFFFF"); ws.cell(r,1).fill=fill(TEAL)
for i,s in enumerate(["Bear","Base","Bull"]):
    c=ws.cell(r,2+i,s); c.font=F(10,True,"FFFFFF"); c.fill=fill([RED_TXT,TEAL,"1E8449"][i]); c.alignment=CEN
r+=1
def srow(label,fmt,fns,bold=False,txt=BLACK):
    global r
    ws.cell(r,1,label).font=F(10,bold)
    for i,fn in enumerate(fns):
        c=ws.cell(r,2+i,fn); c.number_format=fmt; c.font=F(10,bold,txt); c.alignment=CEN
    rr=r; r+=1; return rr
w=R['wacc']
pv=[ "=B{f}/(1+$B${w})^1+C{f}/(1+$B${w})^2+D{f}/(1+$B${w})^3+E{f}/(1+$B${w})^4+F{f}/(1+$B${w})^5".format(f=fcfrow[s],w=w) for s in ["Bear","Base","Bull"]]
pv_r=srow("PV of Explicit FCF ($M)",CUR,pv)
tv=[ "=F{f}*{c}{tm}".format(f=fcfrow[s],c="BCD"[i],tm=R['tmult']) for i,s in enumerate(["Bear","Base","Bull"])]
tv_r=srow("Terminal Value ($M)",CUR,tv)
pvtv=[ "={c}{t}/(1+$B${w})^5".format(c="BCD"[i],t=tv_r,w=w) for i in range(3)]
pvtv_r=srow("PV of Terminal Value ($M)",CUR,pvtv)
ev=[ "={c}{p}+{c}{q}".format(c="BCD"[i],p=pv_r,q=pvtv_r) for i in range(3)]
ev_r=srow("Enterprise Value ($M)",CUR,ev,bold=True)
eq=[ "={c}{e}-$B${nd}".format(c="BCD"[i],e=ev_r,nd=R['netdebt']) for i in range(3)]
eq_r=srow("Equity Value ($M)  [EV - Net Debt]",CUR,eq)
ips=[ "={c}{e}/$B${sh}".format(c="BCD"[i],e=eq_r,sh=R['shares']) for i in range(3)]
ips_r=srow("Intrinsic Price / Share ($)",CUR1,ips,bold=True,txt=GRN_TXT)
for i,fc in enumerate([RED,YEL,GRN]): ws.cell(ips_r,2+i).fill=fill(fc)
cur_r=srow("Current Price ($)",CUR1,["=$B${p}".format(p=R['price'])]*3)
srow("Upside / (Downside)",PCT,[ "={c}{i}/{c}{cu}-1".format(c="BCD"[i],i=ips_r,cu=cur_r) for i in range(3)],bold=True)
R['ips_r']=ips_r

r+=1; sub(ws,r,"Helper — FY2026E Base Case (Dashboard links)",8); r+=1
ws.cell(r,1,"FY2026E Revenue ($M)").font=F(10); c=ws.cell(r,2,"=B{}".format(revrow['Base'])); c.number_format=CUR; c.font=F(10,False,GRN_TXT); R['fy26rev']=r; r+=1
ws.cell(r,1,"FY2026E FCF ($M)").font=F(10); c=ws.cell(r,2,"=B{}".format(fcfrow['Base'])); c.number_format=CUR; c.font=F(10); R['fy26fcf']=r; r+=1
ws.cell(r,1,"FY2026E FCF Margin").font=F(10); c=ws.cell(r,2,"=B{f}/B{rv}".format(f=R['fy26fcf'],rv=R['fy26rev'])); c.number_format=PCT; c.font=F(10); R['fy26fmar']=r; r+=1

# ===================================================== DASHBOARD
ws=wb.create_sheet("Dashboard")
hdr(ws,"TRMB — INVESTMENT DASHBOARD",6)
for col,w in zip("ABCDEF",[36,15,15,15,15,15]): ws.column_dimensions[col].width=w
D={}; r=4
sub(ws,r,"Snapshot",6); r+=1
def drow(key,label,formula,fmt,col=BLACK):
    global r
    ws.cell(r,1,label).font=F(10)
    c=ws.cell(r,2,formula); c.font=F(10,False,col); c.number_format=fmt; D[key]=r; r+=1
drow("price","Current Price","=Model!B{}".format(R['price']),CUR1,BLUE_TXT)
drow("sh","Diluted Shares (M)","=Model!B{}".format(R['shares']),'#,##0',BLUE_TXT)
drow("mc","Market Cap ($M)","=B{p}*B{s}".format(p=D['price'],s=D['sh']),CUR)
drow("cash","Cash & Equivalents ($M)","=Model!B{}".format(R['cash']),CUR,BLUE_TXT)
drow("debt","Total Debt ($M)","=Model!B{}".format(R['debt']),CUR,BLUE_TXT)
drow("nd","Net Debt ($M)","=B{d}-B{c}".format(d=D['debt'],c=D['cash']),CUR)
drow("ev","Enterprise Value ($M)","=B{m}+B{n}".format(m=D['mc'],n=D['nd']),CUR)
r+=1; sub(ws,r,"Operating Metrics (Q1 2026)",6); r+=1
ops=[("Q1'26 Revenue ($M)",939.9,CUR,"TRMB Q1 2026 8-K"),
 ("Q1'26 Revenue Growth YoY",0.12,PCT,"+12% reported & organic"),
 ("ARR ($B)",2.435,'$#,##0.00"B"',"+12% YoY, +12% organic"),
 ("Recurring/Software % of Rev",0.78,PCT,"78% of Q1'26 revenue"),
 ("Non-GAAP Gross Margin",0.71,PCT,"$666.9M"),
 ("Non-GAAP Operating Margin",0.259,PCT,"$243.2M"),
 ("Adj. EBITDA Margin",0.274,PCT,"$257.7M"),
 ("FY2026E Non-GAAP EPS (mid)",3.555,CUR1,"Guidance $3.47-3.64")]
for label,v,fmt,note in ops:
    ws.cell(r,1,label).font=F(10)
    c=ws.cell(r,2,v); c.font=F(10,False,BLUE_TXT); c.number_format=fmt; c.comment=Comment(note,"Model")
    if label.startswith("FY2026E Non-GAAP EPS"): D['eps']=r
    r+=1
r+=1; sub(ws,r,"Valuation Multiples",6); r+=1
def vrow(label,formula,fmt):
    global r
    ws.cell(r,1,label).font=F(10); c=ws.cell(r,2,formula); c.font=F(10); c.number_format=fmt; r+=1
vrow("EV / FY2026E Revenue","=B{ev}/Model!B{rv}".format(ev=D['ev'],rv=R['fy26rev']),MULT)
vrow("Fwd P/E (Price / FY26E NG-EPS)","=B{p}/B{e}".format(p=D['price'],e=D['eps']),MULT)
vrow("EV / FY2026E Adj. EBITDA","=B{ev}/(Model!B{rv}*0.274)".format(ev=D['ev'],rv=R['fy26rev']),MULT)
vrow("FCF Yield (FY26E FCF / Mkt Cap)","=Model!B{f}/B{m}".format(f=R['fy26fcf'],m=D['mc']),PCT)
r+=1; sub(ws,r,"Rule of 40 (hybrid SW/HW)",6); r+=1
ws.cell(r,1,"Organic Growth + Non-GAAP Op Margin").font=F(10); c=ws.cell(r,2,"=0.12+0.259"); c.font=F(10,True,TEAL); c.number_format=PCT; r+=1
ws.cell(r,1,"Organic Growth + FCF Margin (FY26E)").font=F(10); c=ws.cell(r,2,"=0.10+Model!B{}".format(R['fy26fmar'])); c.font=F(10,True,TEAL); c.number_format=PCT; r+=1
r+=1; sub(ws,r,"DCF Intrinsic Value / Share (linked from Model)",6); r+=1
ws.cell(r,1,"Scenario").font=F(10,True,"FFFFFF"); ws.cell(r,1).fill=fill(NAVY)
for i,s in enumerate(["Bear","Base","Bull"]):
    c=ws.cell(r,2+i,s); c.font=F(10,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=CEN
r+=1
ws.cell(r,1,"Intrinsic Price/Share").font=F(10)
for i in range(3):
    c=ws.cell(r,2+i,"=Model!{c}{rr}".format(c="BCD"[i],rr=R['ips_r'])); c.font=F(10,True,GRN_TXT); c.number_format=CUR1; c.alignment=CEN
ipsd=r; r+=1
ws.cell(r,1,"Upside / (Downside) vs. Current").font=F(10)
for i in range(3):
    c=ws.cell(r,2+i,"={c}{ip}/$B${p}-1".format(c="BCD"[i],ip=ipsd,p=D['price'])); c.font=F(10); c.number_format=PCT; c.alignment=CEN
r+=1
r+=1; sub(ws,r,"Technical Levels",6); r+=1
for label,v,fmt in [("52-Week High",87.50,CUR1),("52-Week Low",49.43,CUR1),("50-Day SMA (approx)",62.0,CUR1),
   ("200-Day SMA (approx)",73.0,CUR1),("RSI (14)",26,'0'),("IV Rank (elevated)",0.72,PCT)]:
    ws.cell(r,1,label).font=F(10); c=ws.cell(r,2,v); c.font=F(10,False,BLUE_TXT); c.number_format=fmt; r+=1
ws.cell(r,1,"Price below both SMAs — downtrend; RSI oversold (~26).").font=F(8,False,"808080",True)

# ===================================================== RULE OF 40
ws=wb.create_sheet("Rule_of_40")
hdr(ws,"TRMB — RULE OF 40 & PEER COMPARABLES",7)
for col,w in zip("ABCDEFG",[28,14,15,13,13,14,24]): ws.column_dimensions[col].width=w
r=4; sub(ws,r,"Rule of 40 Scorecard (hybrid software/hardware)",7); r+=1
ws.cell(r,1,"Basis").font=F(10,True,"FFFFFF"); ws.cell(r,1).fill=fill(NAVY)
for i,h in enumerate(["Growth %","Profit %","Rule of 40","Verdict"]):
    c=ws.cell(r,2+i,h); c.font=F(10,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=CEN
r+=1
for label,g,p,v in [("Q1'26 ARR organic + NG Op Margin",0.12,0.259,"Healthy (~38)"),
  ("Q1'26 Rev growth + Adj EBITDA margin",0.12,0.274,"Healthy (~39)"),
  ("FY26E Rev growth + FCF margin",0.10,0.175,"Below 40 (~28)")]:
    ws.cell(r,1,label).font=F(10)
    c=ws.cell(r,2,g); c.font=F(10,False,BLUE_TXT); c.number_format=PCT; c.alignment=CEN
    c=ws.cell(r,3,p); c.font=F(10,False,BLUE_TXT); c.number_format=PCT; c.alignment=CEN
    c=ws.cell(r,4,"=B{r}+C{r}".format(r=r)); c.font=F(10,True,TEAL); c.number_format=PCT; c.alignment=CEN
    ws.cell(r,5,v).font=F(10); r+=1
r+=1; sub(ws,r,"Segment Mix — Q1 2026",7); r+=1
ws.cell(r,1,"Segment").font=F(10,True,"FFFFFF"); ws.cell(r,1).fill=fill(TEAL)
for i,h in enumerate(["Q1'26 Rev ($M)","YoY Growth","% of Total"]):
    c=ws.cell(r,2+i,h); c.font=F(10,True,"FFFFFF"); c.fill=fill(TEAL); c.alignment=CEN
r+=1; ss=r
for name,rev,gr in [("AECO (Arch/Eng/Construction)",391.1,0.166),("Field Systems",409.2,0.139),("Transport & Logistics",139.6,-0.044)]:
    ws.cell(r,1,name).font=F(10)
    c=ws.cell(r,2,rev); c.font=F(10,False,BLUE_TXT); c.number_format=CUR; c.alignment=CEN
    c=ws.cell(r,3,gr); c.font=F(10,False,(GRN_TXT if gr>=0 else RED_TXT)); c.number_format=PCT; c.alignment=CEN
    c=ws.cell(r,4,"=B{r}/SUM($B${a}:$B${b})".format(r=r,a=ss,b=ss+2)); c.font=F(10); c.number_format=PCT; c.alignment=CEN
    r+=1
ws.cell(r,1,"Total").font=F(10,True); c=ws.cell(r,2,"=SUM(B{a}:B{b})".format(a=ss,b=ss+2)); c.font=F(10,True); c.number_format=CUR; c.alignment=CEN
r+=2; sub(ws,r,"Peer Comparables (approximate — relative context only)",7); r+=1
ws.cell(r,1,"Company").font=F(10,True,"FFFFFF"); ws.cell(r,1).fill=fill(NAVY)
for i,h in enumerate(["Ticker","Fwd Rev Gr","FCF Margin","Rule of 40","Fwd EV/Sales","Note"]):
    c=ws.cell(r,2+i,h); c.font=F(10,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=CEN
r+=1
for name,tk,g,f_,evs,note in [("Trimble","TRMB",0.10,0.175,3.3,"Subject — 78% recurring"),
  ("Hexagon AB","HXGBY",0.06,0.20,4.5,"Closest geospatial peer"),
  ("AGCO","AGCO",0.02,0.07,1.1,"Ag/precision, cyclical"),
  ("Topcon","7732.T",0.05,0.08,1.3,"Positioning/medical"),
  ("Bentley Systems","BSY",0.11,0.27,9.5,"Infra software pure-play"),
  ("Autodesk","ADSK",0.11,0.30,8.0,"Design software"),
  ("Zebra Tech","ZBRA",0.06,0.13,2.8,"Enterprise HW+SW")]:
    sub_=tk=="TRMB"
    ws.cell(r,1,name).font=F(10,sub_)
    ws.cell(r,2,tk).font=F(10,sub_); ws.cell(r,2).alignment=CEN
    c=ws.cell(r,3,g); c.font=F(10,False,BLUE_TXT); c.number_format=PCT; c.alignment=CEN
    c=ws.cell(r,4,f_); c.font=F(10,False,BLUE_TXT); c.number_format=PCT; c.alignment=CEN
    c=ws.cell(r,5,"=C{r}+D{r}".format(r=r)); c.font=F(10,sub_,TEAL); c.number_format=PCT; c.alignment=CEN
    c=ws.cell(r,6,evs); c.font=F(10,False,BLUE_TXT); c.number_format=MULT; c.alignment=CEN
    ws.cell(r,7,note).font=F(9)
    if sub_:
        for cc in range(1,8): ws.cell(r,cc).fill=fill(YEL)
    r+=1
ws.cell(r,1,"TRMB trades at a discount to software-pure peers (BSY/ADSK) despite 78% recurring revenue; closer to hardware-blend peers.").font=F(8,False,"808080",True)

# ===================================================== OPTIONS
ws=wb.create_sheet("Options_Strategy")
hdr(ws,"TRMB — OPTIONS STRATEGY OVERLAY (Long-Equity Initiation)",6)
for col,w in zip("ABCDEF",[26,20,16,15,15,30]): ws.column_dimensions[col].width=w
r=4; sub(ws,r,"Market Reference",6); r+=1
for label,v,fmt in [("Current Price",50.00,CUR1),("DCF Base Intrinsic",72.0,CUR1),("DCF Bear",36.0,CUR1),
   ("52-wk Low / High","$49.43 / $87.50",None),("RSI (14)",26,'0'),("IV Rank (elevated)",0.72,PCT)]:
    ws.cell(r,1,label).font=F(10); c=ws.cell(r,2,v); c.font=F(10,False,BLUE_TXT)
    if fmt: c.number_format=fmt
    r+=1
r+=1; sub(ws,r,"Recommended Structures (elevated IV → favor SELLING premium to enter)",6); r+=1
ws.cell(r,1,"Strategy").font=F(10,True,"FFFFFF"); ws.cell(r,1).fill=fill(NAVY)
for i,h in enumerate(["Structure","Est. Prem/Debit","Max Gain","Max Loss","Rationale"]):
    c=ws.cell(r,2+i,h); c.font=F(10,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=CEN
r+=1
for name,struct,prem,mg,ml,rat in [
  ("Cash-Secured Put (initiate)","Sell $45 put, ~60 DTE","~$2.00 cr ($200)","$200 (4.4%/cycle)","$4,300 if→0","High IV pays you to wait; net entry ~$43 vs DCF base $72"),
  ("CSP — deeper/safer","Sell $42.50 put, ~60 DTE","~$1.40 cr ($140)","$140 (3.3%/cycle)","$4,110","Below 52-wk low entry; collect premium during downtrend"),
  ("Long LEAP (leveraged)","Buy $50 call, Jan-2027","~$6.50 debit ($650)","Unbounded","$650","Multi-quarter re-rate to DCF base $72; time the thesis"),
  ("Covered Call (post-assign)","Sell $60 call, ~45 DTE","prem + to $60","capped @ $60","own shares","Harvest IV above cost once long shares")]:
    ws.cell(r,1,name).font=F(10,True)
    for ci,val in enumerate([struct,prem,mg,ml]): ws.cell(r,2+ci,val).font=F(10)
    ws.cell(r,6,rat).font=F(9)
    for cc in range(1,7): ws.cell(r,cc).alignment=Alignment(vertical="center",wrap_text=True)
    ws.row_dimensions[r].height=30; r+=1
r+=1
n=ws.cell(r,1,"Per-cycle yields, NOT annualized. Size so the bear-case ($36) outcome does not cause unacceptable portfolio loss. Not financial advice.")
n.font=F(9,True,RED_TXT); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)

# reorder: Dashboard first
wb.move_sheet("Dashboard",-(wb.sheetnames.index("Dashboard")))
wb.save("TRMB_Investment_Model.xlsx")
print("Saved. ips_r=",R['ips_r'],"order=",wb.sheetnames)
