#!/usr/bin/env python3
"""FSLY (Fastly) investment model — Dashboard, Model/DCF, Rule_of_40, Options_Strategy."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

BLUE = Font(name="Arial", size=10, color="0000FF")          # hardcoded inputs
BLACK = Font(name="Arial", size=10, color="000000")         # formulas
GREEN = Font(name="Arial", size=10, color="008000")         # cross-sheet links
HDR = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=14, bold=True, color="1F4E79")
SUB = Font(name="Arial", size=10, bold=True, color="1F4E79")
BOLD = Font(name="Arial", size=10, bold=True, color="000000")
NAVY = PatternFill("solid", fgColor="1F4E79")
LBLUE = PatternFill("solid", fgColor="D5E8F0")
LGREEN = PatternFill("solid", fgColor="C6EFCE")
LYEL = PatternFill("solid", fgColor="FFEB9C")
LRED = PatternFill("solid", fgColor="FFC7CE")
GREY = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CUR = '$#,##0;($#,##0);-'
CUR2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%;(0.0%);-'
MULT = '0.0x'

wb = Workbook()

def style_block(ws, cell, font=BLACK, fill=None, fmt=None, align=None, border=True):
    c = ws[cell]
    c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align, vertical="center")
    if border: c.border = BORD

# ----------------------------------------------------------------------------
# SHEET 1: DASHBOARD
# ----------------------------------------------------------------------------
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False
for col, w in {"A":2,"B":30,"C":16,"D":16,"E":16,"F":16,"G":16}.items():
    ws.column_dimensions[col].width = w

ws["B2"] = "FASTLY, INC. (NASDAQ: FSLY) — INVESTMENT DASHBOARD"; ws["B2"].font = TITLE
ws["B3"] = "Analysis date: June 8, 2026  |  Analyst: Ed  |  Rating: HOLD / Accumulate on weakness  |  Vehicle: Long equity + income options overlay"
ws["B3"].font = Font(name="Arial", size=9, italic=True, color="595959")

r = 5
ws[f"B{r}"] = "MARKET SNAPSHOT"; ws[f"B{r}"].font = HDR; ws[f"B{r}"].fill = NAVY
for c in ["C","D","E","F","G"]:
    ws[f"{c}{r}"].fill = NAVY
snap = [
    ("Current Price", 20.09, CUR2, "Source: Yahoo Finance / MarketBeat, 6/5-6/7/2026"),
    ("52-Week Range", "$6.29 – $34.82", None, "Source: MarketBeat, 6/2026"),
    ("Shares Outstanding (Class A, M)", 156.4, '#,##0.0', "Source: FSLY 10-Q Q1 2026 cover, 156,367,942 shares"),
    ("Diluted Shares — valuation (M)", 165.0, '#,##0.0', "Analyst est.: basic 156M + dilutive equity awards; converts flagged as overhang"),
    ("Market Cap ($M)", "=C6*C8", CUR, None),
    ("Cash + Marketable Securities ($M)", 330.489, CUR, "Source: FSLY 10-Q Q1 2026 balance sheet (cash 146.670 + securities 183.819)"),
    ("Total Debt — convertible notes ($M)", 323.620, CUR, "Source: FSLY 10-Q Q1 2026, long-term debt net"),
    ("Net Cash ($M)", "=C11-C12", CUR, None),
    ("Enterprise Value ($M)", "=C10-C13", CUR, None),
]
r = 6
for label, val, fmt, note in snap:
    ws[f"B{r}"] = label; ws[f"B{r}"].font = BOLD; ws[f"B{r}"].border = BORD; ws[f"B{r}"].fill = GREY
    cell = ws[f"C{r}"]
    cell.value = val
    if isinstance(val, str) and val.startswith("="):
        cell.font = BLACK
    elif isinstance(val, (int, float)):
        cell.font = BLUE
    else:
        cell.font = BLUE
    if fmt: cell.number_format = fmt
    cell.border = BORD
    cell.alignment = Alignment(horizontal="right")
    if note: cell.comment = Comment(note, "Ed")
    r += 1

r += 1
ws[f"B{r}"] = "FUNDAMENTALS (Most Recent Quarter — Q1 2026)"; ws[f"B{r}"].font = HDR; ws[f"B{r}"].fill = NAVY
for c in ["C","D","E","F","G"]: ws[f"{c}{r}"].fill = NAVY
fund_start = r + 1
fund = [
    ("Q1 2026 Revenue ($M)", 173.021, CUR, "Source: FSLY 8-K Q1 2026 investor supplement"),
    ("Revenue Growth (YoY)", 0.20, PCT, "vs Q1 2025 $144.474M"),
    ("Security Revenue Growth (YoY)", 0.47, PCT, "Security $38.8M = 22% of total revenue"),
    ("GAAP Gross Margin", 0.625, PCT, "Source: FSLY 8-K Q1 2026"),
    ("Non-GAAP Gross Margin", 0.651, PCT, "Record; Source: FSLY 8-K Q1 2026"),
    ("Non-GAAP Operating Margin", 0.111, PCT, "Non-GAAP op income $19.143M / rev"),
    ("Adj. EBITDA Margin", 0.170, PCT, "Adj EBITDA $29.463M / rev"),
    ("FCF Margin (Q1 — heavy capex)", 0.024, PCT, "FCF $4.109M; Q1 capex elevated ($21M PP&E)"),
    ("LTM Net Retention Rate", 1.13, PCT, "Up from 110% Q4 2025, 100% Q1 2025"),
    ("Remaining Perf. Obligations ($M)", 368.7, CUR, "+63% YoY; Source: FSLY 8-K Q1 2026"),
    ("Large Customers (>$100K)", 634, '#,##0', "Source: FSLY 8-K Q1 2026"),
]
r = fund_start
for label, val, fmt, note in fund:
    ws[f"B{r}"] = label; ws[f"B{r}"].font = BOLD; ws[f"B{r}"].border = BORD; ws[f"B{r}"].fill = GREY
    cell = ws[f"C{r}"]; cell.value = val; cell.font = BLUE; cell.number_format = fmt
    cell.border = BORD; cell.alignment = Alignment(horizontal="right")
    if note: cell.comment = Comment(note, "Ed")
    r += 1

# DCF intrinsic values (links from Model sheet) + Rule of 40
dcf_r = 6
ws[f"E{dcf_r}"] = "VALUATION SUMMARY"; ws[f"E{dcf_r}"].font = HDR; ws[f"E{dcf_r}"].fill = NAVY
for c in ["F","G"]: ws[f"{c}{dcf_r}"].fill = NAVY
vals = [
    ("DCF Bull Case / share", "=Model!C40", CUR2, LGREEN),
    ("DCF Base Case / share", "=Model!D40", CUR2, LYEL),
    ("DCF Bear Case / share", "=Model!E40", CUR2, LRED),
    ("Current Price", "=C6", CUR2, None),
    ("Base Case vs Current", "=F8/F10-1", PCT, None),
    ("FY2026 Guided Revenue ($M)", 717.5, CUR, LBLUE),
    ("FY2026 Rev Growth (guidance)", "=F12/Model!B7-1", PCT, None),
    ("EV / FY26 Revenue", "=C14/F12", MULT, None),
    ("Rule of 40 (FY2025)", "=Rule_of_40!C9", '0.0', None),
    ("Rule of 40 (Q1 2026 ann.)", "=Rule_of_40!C16", '0.0', None),
]
r = dcf_r + 1
for label, val, fmt, fill in vals:
    ws[f"E{r}"] = label; ws[f"E{r}"].font = BOLD; ws[f"E{r}"].border = BORD
    if fill: ws[f"E{r}"].fill = fill
    else: ws[f"E{r}"].fill = GREY
    cell = ws[f"F{r}"]; cell.value = val
    cell.font = GREEN if (isinstance(val,str) and "!" in val) else (BLACK if (isinstance(val,str) and val.startswith("=")) else BLUE)
    cell.number_format = fmt; cell.border = BORD; cell.alignment = Alignment(horizontal="right")
    r += 1

tech_r = r + 1
ws[f"E{tech_r}"] = "TECHNICAL LEVELS"; ws[f"E{tech_r}"].font = HDR; ws[f"E{tech_r}"].fill = NAVY
for c in ["F","G"]: ws[f"{c}{tech_r}"].fill = NAVY
tech = [
    ("20-Day SMA (est.)", 21.20, CUR2, "Estimated; stock recovering off ~$19 low"),
    ("50-Day SMA", 23.34, CUR2, "Source: MarketBeat, 6/2026 — price BELOW (resistance)"),
    ("200-Day SMA", 17.14, CUR2, "Source: MarketBeat, 6/2026 — price ABOVE (support)"),
    ("RSI (14, est.)", 41, '0', "Estimated; ~38% drop from May high then bounce"),
    ("Implied Volatility (est.)", "Elevated", None, "Post 38% one-month drop; favors premium selling"),
]
r = tech_r + 1
for label, val, fmt, note in tech:
    ws[f"E{r}"] = label; ws[f"E{r}"].font = BOLD; ws[f"E{r}"].border = BORD; ws[f"E{r}"].fill = GREY
    cell = ws[f"F{r}"]; cell.value = val; cell.font = BLUE
    if fmt: cell.number_format = fmt
    cell.border = BORD; cell.alignment = Alignment(horizontal="right")
    if note: cell.comment = Comment(note, "Ed")
    r += 1

ws[f"B{r+2}"] = "Color key: blue = hardcoded input | black = formula | green = cross-sheet link. Not investment advice."
ws[f"B{r+2}"].font = Font(name="Arial", size=8, italic=True, color="808080")

# ----------------------------------------------------------------------------
# SHEET 2: MODEL (DCF, 3 scenarios, single WACC)
# ----------------------------------------------------------------------------
ms = wb.create_sheet("Model")
ms.sheet_view.showGridLines = False
for col, w in {"A":34,"B":14,"C":14,"D":14,"E":14,"F":14}.items():
    ms.column_dimensions[col].width = w

ms["A1"] = "FSLY — DCF VALUATION MODEL (5-Year FCF, Gordon Growth Terminal)"; ms["A1"].font = TITLE
ms["A2"] = "SINGLE WACC across all scenarios (mandatory). Only growth, FCF margin & terminal growth vary."
ms["A2"].font = Font(name="Arial", size=9, italic=True, color="C00000")

# Historical block
ms["A4"] = "HISTORICAL & BASE ($M)"; ms["A4"].font = HDR; ms["A4"].fill = NAVY
for c in ["B","C","D","E","F"]: ms[f"{c}4"].fill = NAVY
hist = [
    ("FY2024 Revenue", 543.676, "Source: FSLY 8-K FY2025 release (2/11/2026)"),
    ("FY2025 Revenue", 624.018, "Source: FSLY 8-K FY2025 release; +14.8% YoY"),
    ("FY2025 Non-GAAP Op Income", 22.398, "Sum of 4 quarters non-GAAP op income"),
    ("FY2025 Adj. EBITDA", 77.379, "Sum of 4 quarters adj. EBITDA"),
    ("FY2025 Free Cash Flow", 45.809, "Sum of 4 quarters FCF"),
    ("FY2026 Guided Revenue (Yr1 base)", 717.5, "Guidance midpoint $710-725M; label growth FROM this base"),
]
r = 5
ms[f"A{r}"] = "Metric"; ms[f"A{r}"].font = BOLD; ms[f"A{r}"].fill = GREY; ms[f"A{r}"].border = BORD
ms[f"B{r}"] = "Value"; ms[f"B{r}"].font = BOLD; ms[f"B{r}"].fill = GREY; ms[f"B{r}"].border = BORD; ms[f"B{r}"].alignment = Alignment(horizontal="right")
r = 6
for label, val, note in hist:
    ms[f"A{r}"] = label; ms[f"A{r}"].font = BLACK; ms[f"A{r}"].border = BORD
    cell = ms[f"B{r}"]; cell.value = val; cell.font = BLUE; cell.number_format = CUR
    cell.border = BORD; cell.alignment = Alignment(horizontal="right")
    cell.comment = Comment(note, "Ed")
    r += 1
# B6=FY24, B7=FY25, B8 opinc, B9 ebitda, B10 fcf, B11 guided Yr1

# Assumptions block
a0 = 13
ms[f"A{a0}"] = "ASSUMPTIONS"; ms[f"A{a0}"].font = HDR; ms[f"A{a0}"].fill = NAVY
for c in ["B","C","D","E","F"]: ms[f"{c}{a0}"].fill = NAVY
ms[f"A{a0+1}"] = "";
hdr = ["", "", "Bull", "Base", "Bear"]
for i, h in enumerate(["Assumption","", "Bull","Base","Bear"]):
    pass
ms[f"C{a0+1}"] = "Bull"; ms[f"D{a0+1}"] = "Base"; ms[f"E{a0+1}"] = "Bear"
for c in ["C","D","E"]:
    ms[f"{c}{a0+1}"].font = BOLD; ms[f"{c}{a0+1}"].alignment = Alignment(horizontal="center"); ms[f"{c}{a0+1}"].border = BORD
ms[f"C{a0+1}"].fill = LGREEN; ms[f"D{a0+1}"].fill = LYEL; ms[f"E{a0+1}"].fill = LRED
ms[f"A{a0+1}"] = "Assumption"; ms[f"A{a0+1}"].font = BOLD; ms[f"A{a0+1}"].border = BORD; ms[f"A{a0+1}"].fill = GREY

# rows: WACC (single), then growth yr2-5, fcf margin yr1-5, terminal growth
assum = [
    ("WACC (single, all scenarios)", 0.11, 0.11, 0.11, PCT),
    ("Yr2 Revenue Growth", 0.20, 0.13, 0.09, PCT),
    ("Yr3 Revenue Growth", 0.18, 0.12, 0.07, PCT),
    ("Yr4 Revenue Growth", 0.16, 0.11, 0.06, PCT),
    ("Yr5 Revenue Growth", 0.14, 0.10, 0.05, PCT),
    ("Yr1 FCF Margin", 0.08, 0.07, 0.06, PCT),
    ("Yr2 FCF Margin", 0.13, 0.10, 0.07, PCT),
    ("Yr3 FCF Margin", 0.17, 0.13, 0.09, PCT),
    ("Yr4 FCF Margin", 0.22, 0.16, 0.11, PCT),
    ("Yr5 FCF Margin (terminal)", 0.26, 0.19, 0.13, PCT),
    ("Terminal Growth Rate", 0.045, 0.035, 0.030, PCT),
]
r = a0 + 2
for row in assum:
    label, bull, base, bear, fmt = row
    ms[f"A{r}"] = label; ms[f"A{r}"].font = BLACK; ms[f"A{r}"].border = BORD; ms[f"A{r}"].fill = GREY
    for col, v in zip(["C","D","E"], [bull, base, bear]):
        cell = ms[f"{col}{r}"]; cell.value = v; cell.font = BLUE; cell.number_format = fmt
        cell.border = BORD; cell.alignment = Alignment(horizontal="right")
    r += 1
# Assumption row map:
WACC_R = a0+2          # 15
G2,G3,G4,G5 = a0+3, a0+4, a0+5, a0+6   # 16,17,18,19
M1,M2,M3,M4,M5 = a0+7, a0+8, a0+9, a0+10, a0+11  # 20..24
TG = a0+12             # 25

# Revenue projection block
p0 = TG + 2   # 27
ms[f"A{p0}"] = "PROJECTIONS — REVENUE ($M)"; ms[f"A{p0}"].font = HDR; ms[f"A{p0}"].fill = NAVY
for c in ["B","C","D","E","F"]: ms[f"{c}{p0}"].fill = NAVY
ms[f"A{p0+1}"] = "Year"; ms[f"B{p0+1}"]="Bull"; ms[f"C{p0+1}"]="Base"; ms[f"D{p0+1}"]="Bear"
for c in ["A","B","C","D"]:
    ms[f"{c}{p0+1}"].font = BOLD; ms[f"{c}{p0+1}"].border = BORD; ms[f"{c}{p0+1}"].fill = GREY
    ms[f"{c}{p0+1}"].alignment = Alignment(horizontal="center")
# Yr1 (all use guided base B11)
yr1 = p0+2
ms[f"A{yr1}"] = "Yr1 (FY2026, guided)"
for c in ["B","C","D"]:
    ms[f"{c}{yr1}"] = "=$B$11"; ms[f"{c}{yr1}"].font = GREEN; ms[f"{c}{yr1}"].number_format = CUR; ms[f"{c}{yr1}"].border = BORD; ms[f"{c}{yr1}"].alignment = Alignment(horizontal="right")
ms[f"A{yr1}"].font = BLACK; ms[f"A{yr1}"].border = BORD
# Yr2-5 grow from prior
grow_map = [(yr1+1,"Yr2 (FY2027)",G2),(yr1+2,"Yr3 (FY2028)",G3),(yr1+3,"Yr4 (FY2029)",G4),(yr1+4,"Yr5 (FY2030)",G5)]
for rr, label, grow in grow_map:
    ms[f"A{rr}"] = label; ms[f"A{rr}"].font = BLACK; ms[f"A{rr}"].border = BORD
    for col in ["B","C","D"]:
        ms[f"{col}{rr}"] = f"={col}{rr-1}*(1+${col[0] if False else ''}{ {'B':'C','C':'D','D':'E'}[col] }${grow})"
# fix formula references for growth assumptions (Bull col uses C, Base uses D, Bear uses E in assumptions)
col_assum = {"B":"C","C":"D","D":"E"}
for rr, label, grow in grow_map:
    for col in ["B","C","D"]:
        ms[f"{col}{rr}"] = f"={col}{rr-1}*(1+${col_assum[col]}${grow})"
        ms[f"{col}{rr}"].font = BLACK; ms[f"{col}{rr}"].number_format = CUR; ms[f"{col}{rr}"].border = BORD; ms[f"{col}{rr}"].alignment = Alignment(horizontal="right")
rev_rows = {"Yr1":yr1, "Yr2":yr1+1, "Yr3":yr1+2, "Yr4":yr1+3, "Yr5":yr1+4}

# FCF projection block
f0 = yr1+6   # after rev block
ms[f"A{f0}"] = "PROJECTIONS — FREE CASH FLOW ($M)"; ms[f"A{f0}"].font = HDR; ms[f"A{f0}"].fill = NAVY
for c in ["B","C","D","E","F"]: ms[f"{c}{f0}"].fill = NAVY
ms[f"A{f0+1}"]="Year"; ms[f"B{f0+1}"]="Bull"; ms[f"C{f0+1}"]="Base"; ms[f"D{f0+1}"]="Bear"
for c in ["A","B","C","D"]:
    ms[f"{c}{f0+1}"].font = BOLD; ms[f"{c}{f0+1}"].border=BORD; ms[f"{c}{f0+1}"].fill=GREY; ms[f"{c}{f0+1}"].alignment=Alignment(horizontal="center")
margin_rows = {1:M1,2:M2,3:M3,4:M4,5:M5}
fcf_first = f0+2
for i, (yk, rrev) in enumerate([("Yr1",rev_rows["Yr1"]),("Yr2",rev_rows["Yr2"]),("Yr3",rev_rows["Yr3"]),("Yr4",rev_rows["Yr4"]),("Yr5",rev_rows["Yr5"])], start=1):
    rr = fcf_first + (i-1)
    ms[f"A{rr}"] = f"{yk} FCF"; ms[f"A{rr}"].font = BLACK; ms[f"A{rr}"].border = BORD
    mrow = margin_rows[i]
    for col in ["B","C","D"]:
        ms[f"{col}{rr}"] = f"={col}{rrev}*${col_assum[col]}${mrow}"
        ms[f"{col}{rr}"].font = BLACK; ms[f"{col}{rr}"].number_format = CUR; ms[f"{col}{rr}"].border = BORD; ms[f"{col}{rr}"].alignment = Alignment(horizontal="right")
fcf_rows = {f"Yr{i}": fcf_first+(i-1) for i in range(1,6)}

# DCF waterfall
d0 = fcf_first + 6
ms[f"A{d0}"] = "DCF WATERFALL"; ms[f"A{d0}"].font = HDR; ms[f"A{d0}"].fill = NAVY
for c in ["B","C","D","E","F"]: ms[f"{c}{d0}"].fill = NAVY
ms[f"A{d0+1}"]="Item"; ms[f"B{d0+1}"]="Bull"; ms[f"C{d0+1}"]="Base"; ms[f"D{d0+1}"]="Bear"
for c in ["A","B","C","D"]:
    ms[f"{c}{d0+1}"].font=BOLD; ms[f"{c}{d0+1}"].border=BORD; ms[f"{c}{d0+1}"].fill=GREY; ms[f"{c}{d0+1}"].alignment=Alignment(horizontal="center")

# PV of FCF sum
pv_r = d0+2
ms[f"A{pv_r}"] = "PV of 5-Yr FCF"
for col in ["B","C","D"]:
    ac = col_assum[col]
    parts = "+".join([f"{col}{fcf_rows[f'Yr{i}']}/(1+${ac}${WACC_R})^{i}" for i in range(1,6)])
    ms[f"{col}{pv_r}"] = f"={parts}"
    ms[f"{col}{pv_r}"].font=BLACK; ms[f"{col}{pv_r}"].number_format=CUR; ms[f"{col}{pv_r}"].border=BORD; ms[f"{col}{pv_r}"].alignment=Alignment(horizontal="right")
ms[f"A{pv_r}"].font=BLACK; ms[f"A{pv_r}"].border=BORD

# Terminal value (Gordon Growth): FCF_Yr5*(1+g)/(WACC-g)
tv_r = pv_r+1
ms[f"A{tv_r}"] = "Terminal Value (Gordon Growth)"
for col in ["B","C","D"]:
    ac = col_assum[col]
    ms[f"{col}{tv_r}"] = f"={col}{fcf_rows['Yr5']}*(1+${ac}${TG})/(${ac}${WACC_R}-${ac}${TG})"
    ms[f"{col}{tv_r}"].font=BLACK; ms[f"{col}{tv_r}"].number_format=CUR; ms[f"{col}{tv_r}"].border=BORD; ms[f"{col}{tv_r}"].alignment=Alignment(horizontal="right")
ms[f"A{tv_r}"].font=BLACK; ms[f"A{tv_r}"].border=BORD

# PV of terminal
pvt_r = tv_r+1
ms[f"A{pvt_r}"] = "PV of Terminal Value"
for col in ["B","C","D"]:
    ac = col_assum[col]
    ms[f"{col}{pvt_r}"] = f"={col}{tv_r}/(1+${ac}${WACC_R})^5"
    ms[f"{col}{pvt_r}"].font=BLACK; ms[f"{col}{pvt_r}"].number_format=CUR; ms[f"{col}{pvt_r}"].border=BORD; ms[f"{col}{pvt_r}"].alignment=Alignment(horizontal="right")
ms[f"A{pvt_r}"].font=BLACK; ms[f"A{pvt_r}"].border=BORD

# Enterprise value
ev_r = pvt_r+1
ms[f"A{ev_r}"] = "Enterprise Value"
for col in ["B","C","D"]:
    ms[f"{col}{ev_r}"] = f"={col}{pv_r}+{col}{pvt_r}"
    ms[f"{col}{ev_r}"].font=BLACK; ms[f"{col}{ev_r}"].number_format=CUR; ms[f"{col}{ev_r}"].border=BORD; ms[f"{col}{ev_r}"].alignment=Alignment(horizontal="right")
ms[f"A{ev_r}"].font=BOLD; ms[f"A{ev_r}"].border=BORD

# Net cash
nc_r = ev_r+1
ms[f"A{nc_r}"] = "+ Net Cash"
for col in ["B","C","D"]:
    ms[f"{col}{nc_r}"] = "=Dashboard!$C$13"
    ms[f"{col}{nc_r}"].font=GREEN; ms[f"{col}{nc_r}"].number_format=CUR; ms[f"{col}{nc_r}"].border=BORD; ms[f"{col}{nc_r}"].alignment=Alignment(horizontal="right")
ms[f"A{nc_r}"].font=BLACK; ms[f"A{nc_r}"].border=BORD

# Equity value
eq_r = nc_r+1
ms[f"A{eq_r}"] = "Equity Value"
for col in ["B","C","D"]:
    ms[f"{col}{eq_r}"] = f"={col}{ev_r}+{col}{nc_r}"
    ms[f"{col}{eq_r}"].font=BLACK; ms[f"{col}{eq_r}"].number_format=CUR; ms[f"{col}{eq_r}"].border=BORD; ms[f"{col}{eq_r}"].alignment=Alignment(horizontal="right")
ms[f"A{eq_r}"].font=BOLD; ms[f"A{eq_r}"].border=BORD

# Diluted shares
sh_r = eq_r+1
ms[f"A{sh_r}"] = "Diluted Shares (M)"
for col in ["B","C","D"]:
    ms[f"{col}{sh_r}"] = "=Dashboard!$C$9"
    ms[f"{col}{sh_r}"].font=GREEN; ms[f"{col}{sh_r}"].number_format='#,##0.0'; ms[f"{col}{sh_r}"].border=BORD; ms[f"{col}{sh_r}"].alignment=Alignment(horizontal="right")
ms[f"A{sh_r}"].font=BLACK; ms[f"A{sh_r}"].border=BORD

# Intrinsic price per share -> place at row 40 cols C/D/E for Dashboard links? Dashboard expects Model!C40 bull, D40 base, E40 bear.
ips_r = sh_r+1
ms[f"A{ips_r}"] = "INTRINSIC PRICE / SHARE"
# Compute into B/C/D then mirror to C40/D40/E40 to satisfy Dashboard links
for col in ["B","C","D"]:
    ms[f"{col}{ips_r}"] = f"={col}{eq_r}/{col}{sh_r}"
    ms[f"{col}{ips_r}"].number_format=CUR2; ms[f"{col}{ips_r}"].border=BORD; ms[f"{col}{ips_r}"].alignment=Alignment(horizontal="right"); ms[f"{col}{ips_r}"].font=BOLD
ms[f"A{ips_r}"].font=BOLD; ms[f"A{ips_r}"].border=BORD
ms[f"B{ips_r}"].fill = LGREEN; ms[f"C{ips_r}"].fill = LYEL; ms[f"D{ips_r}"].fill = LRED

# Dashboard references Model!C40 (bull), D40 (base), E40 (bear). Map: build explicit row 40.
ms["A40"] = "Dashboard link row (Bull/Base/Bear)"; ms["A40"].font = Font(name="Arial", size=8, italic=True, color="808080")
ms["C40"] = f"=B{ips_r}"; ms["D40"] = f"=C{ips_r}"; ms["E40"] = f"=D{ips_r}"
for c in ["C","D","E"]:
    ms[f"{c}40"].font = BLACK; ms[f"{c}40"].number_format = CUR2

# Premium/discount vs current
pd_r = ips_r+1
ms[f"A{pd_r}"] = "Upside/(Downside) vs Current"
for col, ac in zip(["B","C","D"], ["C","D","E"]):
    ms[f"{col}{pd_r}"] = f"={col}{ips_r}/Dashboard!$C$6-1"
    ms[f"{col}{pd_r}"].font=BLACK; ms[f"{col}{pd_r}"].number_format=PCT; ms[f"{col}{pd_r}"].border=BORD; ms[f"{col}{pd_r}"].alignment=Alignment(horizontal="right")
ms[f"A{pd_r}"].font=BLACK; ms[f"A{pd_r}"].border=BORD

# ----------------------------------------------------------------------------
# SHEET 3: RULE_OF_40
# ----------------------------------------------------------------------------
rs = wb.create_sheet("Rule_of_40")
rs.sheet_view.showGridLines = False
for col, w in {"A":30,"B":14,"C":14,"D":16,"E":16,"F":16,"G":22}.items():
    rs.column_dimensions[col].width = w
rs["A1"] = "FSLY — RULE OF 40 ANALYSIS"; rs["A1"].font = TITLE
rs["A2"] = "Rule of 40 = Revenue Growth % + Profitability Margin %.  >40 healthy, >60 exceptional."
rs["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")

rs["A4"] = "FSLY SCORECARD"; rs["A4"].font = HDR; rs["A4"].fill = NAVY
for c in ["B","C","D"]: rs[f"{c}4"].fill = NAVY
rs["A5"]="Basis"; rs["B5"]="Growth %"; rs["C5"]="Margin %"; rs["D5"]="Rule of 40"
for c in ["A","B","C","D"]:
    rs[f"{c}5"].font=BOLD; rs[f"{c}5"].border=BORD; rs[f"{c}5"].fill=GREY; rs[f"{c}5"].alignment=Alignment(horizontal="center")
# FY2025: growth 14.8%, margins FCF 7.3%, adj EBITDA 12.4%
r40 = [
    ("FY2025 — FCF margin", 0.148, 0.073),
    ("FY2025 — Adj. EBITDA margin", 0.148, 0.124),
    ("FY2025 — Non-GAAP op margin", 0.148, 0.036),
]
r = 6
for label, g, m in r40:
    rs[f"A{r}"]=label; rs[f"A{r}"].font=BLACK; rs[f"A{r}"].border=BORD
    rs[f"B{r}"]=g; rs[f"B{r}"].font=BLUE; rs[f"B{r}"].number_format=PCT; rs[f"B{r}"].border=BORD
    rs[f"C{r}"]=m; rs[f"C{r}"].font=BLUE; rs[f"C{r}"].number_format=PCT; rs[f"C{r}"].border=BORD
    rs[f"D{r}"]=f"=(B{r}+C{r})*100"; rs[f"D{r}"].font=BLACK; rs[f"D{r}"].number_format='0.0'; rs[f"D{r}"].border=BORD
    r += 1
# C9 used by dashboard = FY2025 adj EBITDA-based? Dashboard refs Rule_of_40!C9 for FY2025 R40 and C16 Q1 ann.
# Set D rows are the scores. Dashboard expects a score. Put the headline FY2025 score (adj EBITDA basis) in C9.
rs["A9"]="FY2025 headline R40 (adj. EBITDA basis)"; rs["A9"].font=BOLD; rs["A9"].border=BORD; rs["A9"].fill=GREY
rs["C9"]="=D7"; rs["C9"].font=BLACK; rs["C9"].number_format='0.0'; rs["C9"].border=BORD; rs["C9"].fill=LYEL

rs["A11"]="Q1 2026 (ANNUALIZED YoY)"; rs["A11"].font = HDR; rs["A11"].fill = NAVY
for c in ["B","C","D"]: rs[f"{c}11"].fill = NAVY
rs["A12"]="Basis"; rs["B12"]="Growth %"; rs["C12"]="Margin %"; rs["D12"]="Rule of 40"
for c in ["A","B","C","D"]:
    rs[f"{c}12"].font=BOLD; rs[f"{c}12"].border=BORD; rs[f"{c}12"].fill=GREY; rs[f"{c}12"].alignment=Alignment(horizontal="center")
q1 = [
    ("Q1 2026 — FCF margin", 0.20, 0.024),
    ("Q1 2026 — Adj. EBITDA margin", 0.20, 0.170),
    ("Q1 2026 — Non-GAAP op margin", 0.20, 0.111),
]
r = 13
for label, g, m in q1:
    rs[f"A{r}"]=label; rs[f"A{r}"].font=BLACK; rs[f"A{r}"].border=BORD
    rs[f"B{r}"]=g; rs[f"B{r}"].font=BLUE; rs[f"B{r}"].number_format=PCT; rs[f"B{r}"].border=BORD
    rs[f"C{r}"]=m; rs[f"C{r}"].font=BLUE; rs[f"C{r}"].number_format=PCT; rs[f"C{r}"].border=BORD
    rs[f"D{r}"]=f"=(B{r}+C{r})*100"; rs[f"D{r}"].font=BLACK; rs[f"D{r}"].number_format='0.0'; rs[f"D{r}"].border=BORD
    r += 1
rs["A16"]="Q1 2026 headline R40 (adj. EBITDA basis)"; rs["A16"].font=BOLD; rs["A16"].border=BORD; rs["A16"].fill=GREY
rs["C16"]="=D14"; rs["C16"].font=BLACK; rs["C16"].number_format='0.0'; rs["C16"].border=BORD; rs["C16"].fill=LGREEN

# Peer comps
rs["A19"]="PEER COMPARABLES (CDN / Edge / Security)"; rs["A19"].font = HDR; rs["A19"].fill = NAVY
for c in ["B","C","D","E","F","G"]: rs[f"{c}19"].fill = NAVY
peers_hdr = ["Company","Ticker","Fwd Rev Growth","FCF Margin (est.)","Rule of 40","Fwd EV/Sales","Note"]
for i, h in enumerate(peers_hdr):
    col = chr(ord("A")+i)
    rs[f"{col}20"]=h; rs[f"{col}20"].font=BOLD; rs[f"{col}20"].border=BORD; rs[f"{col}20"].fill=GREY; rs[f"{col}20"].alignment=Alignment(horizontal="center")
peers = [
    ("Fastly","FSLY",0.16,0.10,"=(C21+D21)*100",4.3,"Subject — accelerating Security + AI"),
    ("Cloudflare","NET",0.30,0.12,"=(C22+D22)*100",28.0,"Premium multiple; 30%+ growth"),
    ("Akamai","AKAM",0.06,0.22,"=(C23+D23)*100",4.3,"Mature; security/compute pivot"),
    ("Datadog","DDOG",0.24,0.27,"=(C24+D24)*100",13.0,"Observability; high R40"),
    ("Cloudflare-tier avg","—",0.30,0.12,"=(C25+D25)*100",28.0,"Hypergrowth comp"),
    ("CDN/infra avg","—",0.10,0.18,"=(C26+D26)*100",4.0,"Mature infra comp"),
    ("Limelight/Edgio (defunct)","—","n/a","n/a","n/a","n/a","Cautionary CDN comp"),
]
r = 21
for row in peers:
    name,tk,g,m,r40f,ev,note = row
    rs[f"A{r}"]=name; rs[f"B{r}"]=tk
    rs[f"A{r}"].font = BOLD if name=="Fastly" else BLACK
    for col,val,fmt in [("C",g,PCT),("D",m,PCT),("E",r40f,'0.0'),("F",ev,MULT)]:
        cell=rs[f"{col}{r}"]; cell.value=val
        if isinstance(val,str) and val.startswith("="): cell.font=BLACK
        elif isinstance(val,(int,float)): cell.font=BLUE
        else: cell.font=BLACK
        if fmt and not (isinstance(val,str) and val=="n/a"): cell.number_format=fmt
        cell.border=BORD; cell.alignment=Alignment(horizontal="right")
    rs[f"G{r}"]=note; rs[f"G{r}"].font=Font(name="Arial",size=8); rs[f"G{r}"].border=BORD
    rs[f"B{r}"].font=BLACK; rs[f"B{r}"].border=BORD; rs[f"A{r}"].border=BORD
    if name=="Fastly":
        for col in ["A","B","C","D","E","F","G"]: rs[f"{col}{r}"].fill = LYEL
    r += 1
rs[f"A{r+1}"]="Peer growth/FCF figures are approximate analyst estimates for context, not precise consensus."
rs[f"A{r+1}"].font=Font(name="Arial",size=8,italic=True,color="808080")

# ----------------------------------------------------------------------------
# SHEET 4: OPTIONS_STRATEGY (income overlay on 100-500 shares)
# ----------------------------------------------------------------------------
os_ = wb.create_sheet("Options_Strategy")
os_.sheet_view.showGridLines = False
for col, w in {"A":26,"B":18,"C":16,"D":16,"E":16,"F":30}.items():
    os_.column_dimensions[col].width = w
os_["A1"]="FSLY — OPTIONS OVERLAY (Income; existing 100–500 share position)"; os_["A1"].font=TITLE
os_["A2"]="Goal: generate income against shares. Elevated IV post-drop favors SELLING premium. Yields shown PER CYCLE."
os_["A2"].font=Font(name="Arial",size=9,italic=True,color="595959")

os_["A4"]="MARKET REFERENCE"; os_["A4"].font=HDR; os_["A4"].fill=NAVY
for c in ["B"]: os_[f"{c}4"].fill=NAVY
ref=[("Current Price","=Dashboard!C6",CUR2),("50-Day SMA (resistance)",23.34,CUR2),
     ("200-Day SMA (support)",17.14,CUR2),("DCF Base Case","=Model!D40",CUR2),
     ("DCF Bear Case","=Model!E40",CUR2)]
r=5
for label,val,fmt in ref:
    os_[f"A{r}"]=label; os_[f"A{r}"].font=BOLD; os_[f"A{r}"].border=BORD; os_[f"A{r}"].fill=GREY
    cell=os_[f"B{r}"]; cell.value=val
    cell.font = GREEN if (isinstance(val,str) and "!" in val) else BLUE
    cell.number_format=fmt; cell.border=BORD; cell.alignment=Alignment(horizontal="right")
    r+=1

os_["A11"]="RECOMMENDED STRATEGIES (per 1 contract = 100 shares)"; os_["A11"].font=HDR; os_["A11"].fill=NAVY
for c in ["B","C","D","E","F"]: os_[f"{c}11"].fill=NAVY
cols=["Strategy","Structure","Net Premium / Cost","Max Gain","Max Loss / Risk","Rationale"]
for i,h in enumerate(cols):
    col=chr(ord("A")+i)
    os_[f"{col}12"]=h; os_[f"{col}12"].font=BOLD; os_[f"{col}12"].border=BORD; os_[f"{col}12"].fill=GREY; os_[f"{col}12"].alignment=Alignment(horizontal="center")
strat=[
 ("Covered Call (primary)","Sell $25 call, ~45–60 DTE (Aug 2026)","+$1.00 (~$100/contract)","Premium + $4.91 appreciation to $25 = ~$591/contract","Caps upside above $25; keep shares unless called","Income at/above 50-day SMA resistance. ~5% per-cycle yield on $20 stock. Strike above DCF base, near bull DCF."),
 ("Cash-Secured Put (add)","Sell $17.50 put, ~45 DTE","+$0.70 (~$70/contract)","$70 premium kept if above $17.50","Assigned at $17.50 (net ~$16.80) — $1,680 cash/contract","Adds shares near DCF base/support if dips. ~4% per-cycle yield. Better entry toward intrinsic value."),
 ("Collar (optional)","Sell $25 call + buy $17 put, ~60 DTE","~$0.10 net credit","Premium + appreciation to $25","Downside floored at $17 (near 200-day SMA)","Protects gains given recent 38% drop; near-zero cost. Use if prioritizing protection over pure income."),
]
r=13
for row in strat:
    name,struct,prem,gain,loss,rat=row
    os_[f"A{r}"]=name; os_[f"A{r}"].font=BOLD
    os_[f"B{r}"]=struct; os_[f"C{r}"]=prem; os_[f"D{r}"]=gain; os_[f"E{r}"]=loss; os_[f"F{r}"]=rat
    for col in ["A","B","C","D","E","F"]:
        os_[f"{col}{r}"].border=BORD
        os_[f"{col}{r}"].alignment=Alignment(wrap_text=True, vertical="top")
        if col!="A": os_[f"{col}{r}"].font=Font(name="Arial",size=9)
    os_.row_dimensions[r].height=58
    r+=1

os_[f"A{r+1}"]="Position sizing (100–500 shares): 1–5 covered-call contracts. Size CSP cash so max assignment ($1,680/contract) stays within risk budget."
os_[f"A{r+1}"].font=Font(name="Arial",size=9,italic=True,color="C00000")
os_[f"A{r+2}"]="Premiums are estimates pending live IV; verify on the chain. Yields are PER CYCLE, not annualized. Not investment advice."
os_[f"A{r+2}"].font=Font(name="Arial",size=8,italic=True,color="808080")

import os
out = "/sessions/loving-compassionate-cray/mnt/Stock Ticker Analysis/FSLY/FSLY-2026-06-08/FSLY_Investment_Model.xlsx"
wb.save(out)
print("saved", out)
